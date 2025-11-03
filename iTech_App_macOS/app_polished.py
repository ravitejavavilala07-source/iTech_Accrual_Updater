#!/usr/bin/env python3
"""
iTech Accrual Updater - app_polished.py - FIXED v16
FIXED: File dialog crash on first click
"""

import sys
import os

if 'tensorflow' in sys.modules:
    del sys.modules['tensorflow']
if 'keras' in sys.modules:
    del sys.modules['keras']
if 'torch' in sys.modules:
    del sys.modules['torch']
if 'scipy' in sys.modules:
    del sys.modules['scipy']
if 'sklearn' in sys.modules:
    del sys.modules['sklearn']

import json
import traceback
from pathlib import Path
from datetime import date, datetime
from typing import List, Tuple

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import Qt

try:
    from accrual_updater import AccrualUpdater
except:
    AccrualUpdater = None

try:
    from admin_fee_module_v18b import calculate_admin_fee_for_paysheet
    ADMIN_FEE_AVAILABLE = True
except:
    ADMIN_FEE_AVAILABLE = False

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
CONFIG_FILE = "gui_config.json"
LOGO_PATH = "itech_logo.png"

COLORS = {
    "bg_main": "#2d2d2d",
    "bg_panel": "#3a3a3a",
    "accent": "#0066cc",
    "success": "#34c759",
    "text_main": "#ffffff",
    "text_secondary": "#b0b0b0",
    "border": "#4a4a4a",
}

STYLESHEET = f"""
    QMainWindow {{ background-color: {COLORS["bg_main"]}; }}
    QWidget {{ background-color: {COLORS["bg_main"]}; color: {COLORS["text_main"]}; }}
    QGroupBox {{ border: 1px solid {COLORS["border"]}; border-radius: 6px; margin-top: 10px; padding-top: 10px; font-weight: 600; color: {COLORS["text_main"]}; }}
    QGroupBox::title {{ subcontrol-origin: margin; left: 10px; padding: 0 3px 0 3px; }}
    QLineEdit, QSpinBox {{ background-color: {COLORS["bg_panel"]}; color: {COLORS["text_main"]}; border: 1px solid {COLORS["border"]}; border-radius: 4px; padding: 6px; selection-background-color: {COLORS["accent"]}; }}
    QLineEdit:focus, QSpinBox:focus {{ border: 2px solid {COLORS["accent"]}; }}
    QComboBox {{ background-color: {COLORS["bg_panel"]}; color: {COLORS["text_main"]}; border: 1px solid {COLORS["border"]}; border-radius: 4px; padding: 6px; }}
    QComboBox:focus {{ border: 2px solid {COLORS["accent"]}; }}
    QComboBox QAbstractItemView {{ background-color: {COLORS["bg_panel"]}; color: {COLORS["text_main"]}; selection-background-color: {COLORS["accent"]}; }}
    QPushButton {{ background-color: {COLORS["accent"]}; color: white; border: none; border-radius: 4px; padding: 8px; font-weight: 600; }}
    QPushButton:hover {{ background-color: #0052a3; }}
    QPushButton:pressed {{ background-color: #003d7a; }}
    QPushButton:disabled {{ background-color: #666666; }}
    #runBtn {{ background-color: {COLORS["success"]}; min-height: 40px; font-size: 14px; }}
    #runBtn:hover {{ background-color: #2aa048; }}
    QTableView {{ background-color: {COLORS["bg_panel"]}; gridline-color: {COLORS["border"]}; border: 1px solid {COLORS["border"]}; border-radius: 4px; }}
    QTableView::item {{ padding: 4px; }}
    QTableView::item:selected {{ background-color: {COLORS["accent"]}; }}
    QHeaderView::section {{ background-color: {COLORS["bg_panel"]}; color: {COLORS["text_main"]}; padding: 4px; border: none; border-right: 1px solid {COLORS["border"]}; border-bottom: 1px solid {COLORS["border"]}; }}
    QCheckBox {{ color: {COLORS["text_main"]}; }}
    QCheckBox::indicator {{ width: 18px; height: 18px; }}
    QCheckBox::indicator:unchecked {{ background-color: {COLORS["bg_panel"]}; border: 1px solid {COLORS["border"]}; border-radius: 3px; }}
    QCheckBox::indicator:checked {{ background-color: {COLORS["accent"]}; border: 1px solid {COLORS["accent"]}; border-radius: 3px; }}
    QPlainTextEdit {{ background-color: {COLORS["bg_panel"]}; color: {COLORS["text_main"]}; border: 1px solid {COLORS["border"]}; border-radius: 4px; padding: 4px; }}
    QLabel {{ color: {COLORS["text_main"]}; }}
"""


class CalendarDialog(QtWidgets.QDialog):
    def __init__(self, parent, initial_date):
        super().__init__(parent)
        self.setWindowTitle("Select Date")
        self.setModal(True)
        self.date_value = initial_date
        layout = QtWidgets.QVBoxLayout()
        self.cal = QtWidgets.QCalendarWidget()
        self.cal.setSelectedDate(initial_date)
        layout.addWidget(self.cal)
        btns = QtWidgets.QHBoxLayout()
        ok_btn = QtWidgets.QPushButton("OK")
        cancel_btn = QtWidgets.QPushButton("Cancel")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)
        self.setLayout(layout)
    def get_date(self):
        return self.cal.selectedDate().toPython()


class DateModel(QtCore.QAbstractTableModel):
    def __init__(self):
        super().__init__()
        self.rows = []
    def rowCount(self, parent=None):
        return len(self.rows)
    def columnCount(self, parent=None):
        return 2
    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            d, m = self.rows[index.row()]
            if index.column() == 0:
                return d.strftime("%m/%d/%Y")
            else:
                return f"{m:.2f}"
        return None
    def headerData(self, section, orientation, role):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return ["Date", "Multiplier"][section]
        return None
    def add_row(self, d, m):
        self.beginInsertRows(QtCore.QModelIndex(), len(self.rows), len(self.rows))
        self.rows.append((d, m))
        self.endInsertRows()
    def remove_row(self, idx):
        if 0 <= idx < len(self.rows):
            self.beginRemoveRows(QtCore.QModelIndex(), idx, idx)
            self.rows.pop(idx)
            self.endRemoveRows()
    def clear(self):
        self.beginResetModel()
        self.rows = []
        self.endResetModel()


class RunnerThread(QtCore.QThread):
    log_signal = QtCore.Signal(str)
    done_signal = QtCore.Signal()
    error_signal = QtCore.Signal(str)
    def __init__(self, master, paysheets, month, year, dry_run, enable_accrual, enable_admin_fee, pay_dates):
        super().__init__()
        self.master = master
        self.paysheets = paysheets
        self.month = month
        self.year = year
        self.dry_run = dry_run
        self.enable_accrual = enable_accrual and AccrualUpdater is not None
        self.enable_admin_fee = enable_admin_fee and ADMIN_FEE_AVAILABLE
        self.pay_dates = pay_dates
    def run(self):
        try:
            if self.enable_accrual:
                self.log_signal.emit("\n" + "="*80)
                self.log_signal.emit("STEP 1: ACCRUAL UPDATE (Hours, Billed, Pay Dates)")
                self.log_signal.emit("="*80 + "\n")
                date_multiplier_pairs = []
                for d, m in self.pay_dates:
                    date_str = d.strftime("%m/%d/%y")
                    date_multiplier_pairs.append((date_str, m))
                updater = AccrualUpdater(
                    master_path=self.master,
                    sheet_name="Profit Sharing",
                    header_row=3,
                    month=self.month,
                    year=self.year,
                    paysheets_folder=self.paysheets,
                    date_multiplier_pairs=date_multiplier_pairs,
                    dry_run=self.dry_run,
                    backup=True,
                )
                result = updater.process()
                for line in updater.log_lines:
                    self.log_signal.emit(line)
            if self.enable_admin_fee:
                self.log_signal.emit("\n" + "="*80)
                self.log_signal.emit("STEP 2: ADMIN FEE CALCULATION (Column T)")
                self.log_signal.emit("="*80 + "\n")
                from openpyxl import load_workbook
                import re
                month_idx = MONTHS.index(self.month) + 1
                self.log_signal.emit(f"Loading master: {self.master}")
                wb = load_workbook(self.master)
                ws = wb['Profit Sharing']
                self.log_signal.emit("✓ Loaded\n")
                lookup = {}
                for r in range(4, ws.max_row + 1):
                    fn = ws.cell(row=r, column=1).value
                    if fn:
                        fn_str = str(fn).strip()
                        if len(fn_str) == 6 and fn_str.isdigit():
                            lookup[fn_str] = r
                self.log_signal.emit(f"Found {len(lookup)} file numbers in master\n")
                files = []
                paysheets_path = Path(self.paysheets)
                self.log_signal.emit(f"Searching paysheets folder: {paysheets_path}\n")
                try:
                    for psheet in paysheets_path.rglob('*.xls'):
                        files.append(str(psheet))
                    for psheet in paysheets_path.rglob('*.xlsx'):
                        files.append(str(psheet))
                except Exception as e:
                    self.log_signal.emit(f"❌ Error searching folders: {e}\n")
                files = sorted(files)
                self.log_signal.emit(f"Found {len(files)} paysheets (including subfolders)\n\n")
                if len(files) == 0:
                    self.log_signal.emit("❌ NO PAYSHEETS FOUND!\n")
                    self.done_signal.emit()
                    return
                updated = 0
                skipped = 0
                T_COL = 20
                for idx, psheet in enumerate(files, 1):
                    fname = os.path.basename(psheet)
                    m = re.search(r'_(\d{6})\.xls', fname)
                    if not m:
                        skipped += 1
                        continue
                    file_number = m.group(1)
                    if file_number not in lookup:
                        skipped += 1
                        continue
                    mrow = lookup[file_number]
                    try:
                        hours, rate, admin_fee = calculate_admin_fee_for_paysheet(
                            paysheet_path=psheet,
                            month=month_idx,
                            year=self.year,
                            debug=False
                        )
                        if isinstance(admin_fee, (int, float)) and admin_fee > 0:
                            if not self.dry_run:
                                ws.cell(row=mrow, column=T_COL).value = round(float(admin_fee), 2)
                                updated += 1
                            self.log_signal.emit(f"[{idx:3d}] {fname}: ${admin_fee:.2f}")
                        else:
                            skipped += 1
                    except Exception as e:
                        self.log_signal.emit(f"[{idx:3d}] ✗ {fname}: {str(e)[:60]}")
                if not self.dry_run:
                    try:
                        wb.save(self.master)
                        self.log_signal.emit(f"✓ Saved! Updated {updated}/{len(files)} admin fees\n")
                    except Exception as e:
                        self.log_signal.emit(f"✗ Error saving: {e}\n")
            self.done_signal.emit()
        except Exception as e:
            self.log_signal.emit(f"\n✗ ERROR: {e}\n{traceback.format_exc()}")
            self.error_signal.emit(str(e))


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("iTech Accrual Updater")
        self.resize(900, 700)
        self.setStyleSheet(STYLESHEET)
        self.runner = None
        self._build_ui()
        self._load_config()
        QtCore.QTimer.singleShot(500, self._warm_dialogs)

    def _warm_dialogs(self):
        """Pre-initialize file dialogs to prevent first-click crash"""
        try:
            d = QtWidgets.QFileDialog(self)
            d.setFileMode(QtWidgets.QFileDialog.ExistingFile)
        except:
            pass

    def _build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        main_layout = QtWidgets.QVBoxLayout(central)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(10)
        header_layout = QtWidgets.QHBoxLayout()
        if Path(LOGO_PATH).exists():
            logo_label = QtWidgets.QLabel()
            pixmap = QtGui.QPixmap(LOGO_PATH)
            pixmap = pixmap.scaledToHeight(50, Qt.SmoothTransformation)
            logo_label.setPixmap(pixmap)
            header_layout.addWidget(logo_label)
        title = QtWidgets.QLabel("iTech Accrual Updater")
        title.setFont(QtGui.QFont("Arial", 18, QtGui.QFont.Bold))
        header_layout.addWidget(title)
        header_layout.addStretch()
        main_layout.addLayout(header_layout)
        master_grp = QtWidgets.QGroupBox("Master File")
        master_layout = QtWidgets.QHBoxLayout()
        self.master_input = QtWidgets.QLineEdit()
        self.master_input.setReadOnly(True)
        master_btn = QtWidgets.QPushButton("Browse")
        master_btn.setMaximumWidth(100)
        master_btn.clicked.connect(self._browse_master)
        master_layout.addWidget(self.master_input)
        master_layout.addWidget(master_btn)
        master_grp.setLayout(master_layout)
        main_layout.addWidget(master_grp)
        paysheets_grp = QtWidgets.QGroupBox("Paysheets Folder (searches subfolders)")
        paysheets_layout = QtWidgets.QHBoxLayout()
        self.paysheets_input = QtWidgets.QLineEdit()
        self.paysheets_input.setReadOnly(True)
        paysheets_btn = QtWidgets.QPushButton("Browse")
        paysheets_btn.setMaximumWidth(100)
        paysheets_btn.clicked.connect(self._browse_paysheets)
        paysheets_layout.addWidget(self.paysheets_input)
        paysheets_layout.addWidget(paysheets_btn)
        paysheets_grp.setLayout(paysheets_layout)
        main_layout.addWidget(paysheets_grp)
        date_grp = QtWidgets.QGroupBox("Period")
        date_layout = QtWidgets.QHBoxLayout()
        date_layout.addWidget(QtWidgets.QLabel("Month:"))
        self.month_combo = QtWidgets.QComboBox()
        self.month_combo.addItems(MONTHS)
        date_layout.addWidget(self.month_combo)
        date_layout.addWidget(QtWidgets.QLabel("Year:"))
        self.year_spin = QtWidgets.QSpinBox()
        self.year_spin.setValue(datetime.now().year)
        self.year_spin.setRange(2020, 2030)
        date_layout.addWidget(self.year_spin)
        date_layout.addStretch()
        date_grp.setLayout(date_layout)
        main_layout.addWidget(date_grp)
        pd_grp = QtWidgets.QGroupBox("Pay Date Multipliers (Optional)")
        pd_layout = QtWidgets.QVBoxLayout()
        pd_btn_layout = QtWidgets.QHBoxLayout()
        add_btn = QtWidgets.QPushButton("Add Date")
        add_btn.clicked.connect(self._add_date)
        edit_btn = QtWidgets.QPushButton("Edit Date")
        edit_btn.clicked.connect(self._edit_date)
        del_btn = QtWidgets.QPushButton("Delete Date")
        del_btn.clicked.connect(self._delete_date)
        pd_btn_layout.addWidget(add_btn)
        pd_btn_layout.addWidget(edit_btn)
        pd_btn_layout.addWidget(del_btn)
        pd_btn_layout.addStretch()
        pd_layout.addLayout(pd_btn_layout)
        self.pd_model = DateModel()
        self.pd_view = QtWidgets.QTableView()
        self.pd_view.setModel(self.pd_model)
        self.pd_view.setMaximumHeight(100)
        self.pd_view.horizontalHeader().setStretchLastSection(True)
        pd_layout.addWidget(self.pd_view)
        pd_grp.setLayout(pd_layout)
        main_layout.addWidget(pd_grp)
        opts = QtWidgets.QHBoxLayout()
        self.dry_run = QtWidgets.QCheckBox("Dry Run")
        self.dry_run.setChecked(True)
        opts.addWidget(self.dry_run)
        self.accrual_cb = QtWidgets.QCheckBox("Accrual (Hours/Billed/AB)")
        self.accrual_cb.setChecked(AccrualUpdater is not None)
        self.accrual_cb.setEnabled(AccrualUpdater is not None)
        opts.addWidget(self.accrual_cb)
        self.admin_fee = QtWidgets.QCheckBox("Admin Fee (Column T)")
        self.admin_fee.setChecked(ADMIN_FEE_AVAILABLE)
        self.admin_fee.setEnabled(ADMIN_FEE_AVAILABLE)
        opts.addWidget(self.admin_fee)
        opts.addStretch()
        main_layout.addLayout(opts)
        self.run_btn = QtWidgets.QPushButton("RUN UPDATE")
        self.run_btn.setObjectName("runBtn")
        self.run_btn.setMinimumHeight(35)
        self.run_btn.clicked.connect(self._run)
        main_layout.addWidget(self.run_btn)
        self.log = QtWidgets.QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setFont(QtGui.QFont("Courier", 9))
        main_layout.addWidget(self.log, 1)

    def _browse_master(self):
        try:
            f, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Master File", "", "Excel (*.xlsx *.xls)")
            if f:
                self.master_input.setText(f)
                self._save()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Failed to open file dialog: {e}")

    def _browse_paysheets(self):
        try:
            f = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Paysheets Folder")
            if f:
                self.paysheets_input.setText(f)
                self._save()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Failed to open folder dialog: {e}")

    def _add_date(self):
        cal = CalendarDialog(self, date.today())
        if cal.exec() == QtWidgets.QDialog.Accepted:
            d = cal.get_date()
            m, ok = QtWidgets.QInputDialog.getDouble(self, "", "Multiplier:", 1.0, 0, 10, 2)
            if ok:
                self.pd_model.add_row(d, m)
                self._save()

    def _edit_date(self):
        sel = self.pd_view.selectionModel().selectedRows()
        if not sel:
            return
        idx = sel[0].row()
        old_d, old_m = self.pd_model.rows[idx]
        cal = CalendarDialog(self, old_d)
        if cal.exec() == QtWidgets.QDialog.Accepted:
            d = cal.get_date()
            m, ok = QtWidgets.QInputDialog.getDouble(self, "", "Multiplier:", old_m, 0, 10, 2)
            if ok:
                self.pd_model.remove_row(idx)
                self.pd_model.add_row(d, m)
                self._save()

    def _delete_date(self):
        sel = self.pd_view.selectionModel().selectedRows()
        if sel:
            self.pd_model.remove_row(sel[0].row())
            self._save()

    def _run(self):
        master = self.master_input.text().strip()
        paysheets = self.paysheets_input.text().strip()
        if not master or not Path(master).exists():
            QtWidgets.QMessageBox.critical(self, "Error", "Select valid master file")
            return
        if not paysheets or not Path(paysheets).is_dir():
            QtWidgets.QMessageBox.critical(self, "Error", "Select valid paysheets folder")
            return
        if not self.dry_run.isChecked():
            if QtWidgets.QMessageBox.question(self, "", "Modify master file?") != QtWidgets.QMessageBox.Yes:
                return
        self._save()
        self.run_btn.setEnabled(False)
        self.log.clear()
        self.runner = RunnerThread(
            master, paysheets, self.month_combo.currentText(),
            self.year_spin.value(), self.dry_run.isChecked(),
            self.accrual_cb.isChecked(), self.admin_fee.isChecked(),
            self.pd_model.rows
        )
        self.runner.log_signal.connect(self.log.appendPlainText)
        self.runner.done_signal.connect(lambda: self.run_btn.setEnabled(True))
        self.runner.start()

    def _save(self):
        try:
            cfg = {
                "master": self.master_input.text(),
                "paysheets": self.paysheets_input.text(),
                "month": self.month_combo.currentIndex(),
                "year": self.year_spin.value(),
                "dry_run": self.dry_run.isChecked(),
                "pay_dates": [(d.isoformat(), m) for d, m in self.pd_model.rows],
            }
            with open(CONFIG_FILE, "w") as f:
                json.dump(cfg, f)
        except:
            pass

    def _load_config(self):
        if not Path(CONFIG_FILE).exists():
            return
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
            self.master_input.setText(cfg.get("master", ""))
            self.paysheets_input.setText(cfg.get("paysheets", ""))
            self.month_combo.setCurrentIndex(cfg.get("month", 0))
            self.year_spin.setValue(cfg.get("year", datetime.now().year))
            self.dry_run.setChecked(cfg.get("dry_run", True))
            for d_str, m in cfg.get("pay_dates", []):
                try:
                    self.pd_model.add_row(datetime.fromisoformat(d_str).date(), m)
                except:
                    pass
        except:
            pass


def main():
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
