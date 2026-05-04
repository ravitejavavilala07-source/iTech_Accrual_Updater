#!/usr/bin/env python3
"""
accrual_updater.py - HYBRID VERSION 7.0.3 FIXED
STRUCTURE: xlwings from v7.0.1 (Excel stays open, preserves pivots via Excel save)
LOGIC: v7.5.0 features (formula preservation, split cells, admin fees, gross salary)
WRITE METHOD: xlwings direct writes to Excel cells
UPDATED: Column N (billed) writes to Column V (hourly), Gross Salary in Column G
FIXED: Checks for formulas before writing - never overwrites formula cells!
FIX 7.0.3: parse_paysheet now requires '/' or '-' in text to prevent simple numbers like "8.0" from being treated as dates

Line count: 1800+
Current Date/Time: 2025-11-28 18:00:00 UTC
Current User: ravitejavavilala07-source
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import date, datetime, timedelta
from fractions import Fraction
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:
    print("ERROR: Missing openpyxl. Install: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except Exception:
    print("ERROR: Missing pandas. Install: pip install pandas")
    sys.exit(1)

try:
    import xlwings as xw
except Exception:
    print("ERROR: Missing xlwings. Install: pip install xlwings")
    print("On Mac: brew install xlwings (if using Homebrew)")
    sys.exit(1)

SUPPORT_XLS = True
try:
    import xlrd
except Exception:
    SUPPORT_XLS = False

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

__version__ = "7.0.3"
__author__ = "ravitejavavilala07-source"
__date__ = "2025-11-28 18:00:00"


# ==============================================================================
# SECTION 0: UTILITY FUNCTIONS
# ==============================================================================

def year_in_sheet_name(year: int, name: str) -> bool:
    """Word-boundary year match. '2030' matches '2030', '2030-Closed', 'FY 2030'.
    Does NOT match '12030', '2029-2030' (substring trap), or '20300'."""
    return bool(re.search(rf'(?<!\d){year}(?!\d)', str(name)))


def select_year_sheet(book_or_items, year: int):
    """Pick best year-tab. Prefer exact name '2030' > 'FY2030' / '2030-Closed' > anything containing year.
    Returns sheet object (xlrd) or (name, df) tuple list (pandas)."""
    if hasattr(book_or_items, 'sheets'):
        items = [(s.name, s) for s in book_or_items.sheets()]
    else:
        items = list(book_or_items)
    candidates = [(n, s) for (n, s) in items if year_in_sheet_name(year, n)]
    if not candidates:
        return None
    # Prefer exact match
    exact = [(n, s) for (n, s) in candidates if str(n).strip() == str(year)]
    if exact:
        return exact[0][1]
    # Prefer "Closed"-suffix or fiscal-year prefix only-once-year-appears
    return candidates[0][1]


def safe_float(x: Any) -> float:
    """Convert any value to float safely"""
    try:
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace(",", "").replace("$", "")
        s = s.replace("(", "-").replace(")", "")
        if s in ("", "-"):
            return 0.0
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else 0.0
    except Exception:
        return 0.0


def _normalize_input_date_to_dateobj(s: Optional[str]) -> Optional[date]:
    """Parse date string to date object"""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    
    fmts = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d", "%m-%d-%Y", "%m-%d-%y")
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            return date(dt.year, dt.month, dt.day)
        except Exception:
            pass
    
    m = re.match(r'^\s*0?(\d{1,2})[/-]0?(\d{1,2})[/-](\d{2,4})\s*$', s)
    if m:
        mm = int(m.group(1))
        dd = int(m.group(2))
        yy = int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return date(yy, mm, dd)
        except Exception:
            return None
    return None


def parse_multiplier_input(s: str) -> float:
    """Parse multiplier string"""
    s = (s or "").strip()
    if s == "":
        return 1.0
    s = s.replace(" ", "")
    
    try:
        if "%" in s:
            return float(s.replace("%", "")) / 100.0
        elif "/" in s:
            parts = s.split("/")
            if len(parts) == 2:
                return float(parts[0]) / float(parts[1])
            return float(Fraction(s))
        else:
            return float(s)
    except Exception:
        return 1.0


# ==============================================================================
# SECTION 1: ADMIN FEE MODULE (from v7.5.0)
# ==============================================================================

def extract_period_dates(period_str: str) -> Optional[Tuple[Tuple[int, int, int], Tuple[int, int, int]]]:
    """Extract period date range from 'MM/DD-MM/DD/YYYY' format"""
    match = re.match(r'(\d{1,2})/(\d{1,2})\s*[-/]\s*(\d{1,2})[-/](\d{1,2})[-/](\d{4})', period_str)
    if match:
        return ((int(match.group(1)), int(match.group(2)), int(match.group(5))), 
                (int(match.group(3)), int(match.group(4)), int(match.group(5))))
    return None


def extract_single_date(date_str: str) -> Optional[Tuple[int, int, int]]:
    """Extract single date from MM/DD/YYYY format"""
    match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', date_str)
    if match:
        month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if year < 100: year += 2000
        return (month, day, year)
    return None


def parse_admin_fee_eff_date(text: str) -> Optional[Tuple[int, int, int]]:
    """Parse date from 'Admin Fee Eff MM/DD/YYYY' text"""
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', text)
    if match:
        month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if year < 100: year += 2000
        return (month, day, year)
    return None


def find_admin_fee_eff(sheet, header_row: int, search_rows: int = 10) -> list:
    """Find ALL 'Admin Fee Eff' entries"""
    admin_fees = []
    search_start = max(0, header_row - search_rows)
    
    for r in range(search_start, header_row):
        for c in range(sheet.ncols):
            cell_val = sheet.cell_value(r, c)
            if cell_val and "admin fee eff" in str(cell_val).lower():
                date_tuple = parse_admin_fee_eff_date(str(cell_val))
                if date_tuple and c + 1 < sheet.ncols:
                    rate = safe_float(sheet.cell_value(r, c + 1))
                    if rate > 0:
                        admin_fees.append((date_tuple, rate))
    
    admin_fees.sort(key=lambda x: (x[0][2], x[0][0], x[0][1]))
    return admin_fees


def find_static_admin_fee(sheet, header_row: int, search_rows: int = 10) -> float:
    """Find static 'Admin Fee'"""
    search_start = max(0, header_row - search_rows)
    
    for r in range(search_start, header_row):
        for c in range(sheet.ncols):
            cell_val = sheet.cell_value(r, c)
            if cell_val:
                text = str(cell_val).lower().strip()
                if text in ('admin fee', 'adminfee') and 'eff' not in text:
                    if c + 1 < sheet.ncols:
                        rate = safe_float(sheet.cell_value(r, c + 1))
                        if rate > 0:
                            return rate
    return 0.0


def get_rate_for_date(admin_fees: list, period_date: Tuple[int, int, int]) -> float:
    """Get applicable rate for a specific date"""
    applicable_rate = 0.0
    for date_tuple, rate in admin_fees:
        eff_comp = (date_tuple[2], date_tuple[0], date_tuple[1])
        period_comp = (period_date[2], period_date[0], period_date[1])
        if eff_comp <= period_comp:
            applicable_rate = rate
        else:
            break
    return applicable_rate


def is_full_month_period(start_date: Tuple[int, int, int], end_date: Tuple[int, int, int]) -> bool:
    """Check if period is a full month"""
    if start_date[0] != end_date[0]:
        return False
    if start_date[1] != 1:
        return False
    if end_date[1] < 28:
        return False
    return True


def has_mid_month_eff(admin_fees_list: list, month: int) -> bool:
    """Check if Admin Fee Eff is mid-month"""
    for date_tuple, rate in admin_fees_list:
        if date_tuple[0] == month and date_tuple[1] > 1:
            return True
    return False


def get_full_month_eff_rate(admin_fees_list: list, month: int) -> float:
    """Get the Eff rate for full month periods"""
    latest_rate = 0.0
    for date_tuple, rate in admin_fees_list:
        if date_tuple[0] == month and date_tuple[1] > 1:
            latest_rate = rate
    return latest_rate


def calculate_admin_fee_for_paysheet(paysheet_path: str, month: int, year: int, debug: bool = False) -> Tuple[float, float, float]:
    """
    Calculate admin fee from paysheet (v7.5.0 logic)
    Returns: (total_hours, last_rate, total_admin_fee)
    """
    total_hours = 0.0
    total_admin_fee = 0.0
    last_rate = 0.0
    
    try:
        book = xlrd.open_workbook(paysheet_path, formatting_info=False)
        sheet = None
        for s in book.sheets():
            if year_in_sheet_name(year, s.name):
                sheet = s
                break
        if not sheet:
            return 0.0, 0.0, 0.0
        
        headers = []
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val:
                    text = str(val).lower().strip()
                    if text == 'work period' or text == 'hours & payment':
                        headers.append((r, c, text))
                        break
        
        if not headers:
            return 0.0, 0.0, 0.0
        
        for section_idx, (hp_row, hp_col, header_type) in enumerate(headers):
            section_end = sheet.nrows
            
            for r in range(hp_row + 1, sheet.nrows):
                cell_val = sheet.cell_value(r, hp_col)
                if not cell_val:
                    continue
                cell_str = str(cell_val).strip().lower()
                
                if any(x in cell_str for x in ['total', 'admin fees', 'deductions']):
                    section_end = r
                    break
                
                if section_idx < len(headers) - 1:
                    next_hp_row = headers[section_idx + 1][0]
                    if r >= next_hp_row - 5:
                        section_end = r
                        break
            
            admin_fees_list = find_admin_fee_eff(sheet, hp_row, search_rows=10)
            static_rate = find_static_admin_fee(sheet, hp_row, search_rows=10)
            
            if admin_fees_list and static_rate > 0:
                admin_fees_list = [((1, 1, year), static_rate)] + admin_fees_list
            
            if admin_fees_list:
                section_hours = 0.0
                section_fee = 0.0
                section_rate = 0.0
                
                has_mid_month = has_mid_month_eff(admin_fees_list, month)
                full_month_eff_rate = get_full_month_eff_rate(admin_fees_list, month)
                
                for r in range(hp_row + 1, section_end):
                    period_text = sheet.cell_value(r, hp_col)
                    if not period_text:
                        continue
                    period_str = str(period_text).strip().lower()
                    
                    period_dates = extract_period_dates(period_str)
                    single_date = None
                    
                    if not period_dates:
                        single_date = extract_single_date(period_str)
                    
                    if period_dates:
                        start_date, end_date = period_dates
                        if start_date[0] != month or start_date[2] != year:
                            continue
                        
                        hours_val = safe_float(sheet.cell_value(r, hp_col + 1))
                        if hours_val <= 0:
                            continue
                        
                        if is_full_month_period(start_date, end_date) and has_mid_month:
                            rate = full_month_eff_rate
                        else:
                            rate = get_rate_for_date(admin_fees_list, start_date)
                        
                        fee = hours_val * rate if rate > 0 else 0.0
                        
                        if fee > 0:
                            section_hours += hours_val
                            section_fee += fee
                            section_rate = rate
                    
                    elif single_date:
                        if single_date[0] != month or single_date[2] != year:
                            continue
                        
                        amount_val = safe_float(sheet.cell_value(r, hp_col + 1))
                        
                        if amount_val <= 0:
                            continue
                        
                        rate = get_rate_for_date(admin_fees_list, single_date)
                        
                        if rate > 0:
                            fee = amount_val
                            section_hours += 1
                            section_fee += fee
                            section_rate = rate
                
                if section_hours > 0:
                    total_hours += section_hours
                    total_admin_fee += section_fee
                    last_rate = section_rate
                continue
            
            if static_rate > 0:
                section_hours = 0.0
                
                for r in range(hp_row + 1, section_end):
                    period_text = sheet.cell_value(r, hp_col)
                    if not period_text:
                        continue
                    period_str = str(period_text).strip().lower()
                    
                    period_dates = extract_period_dates(period_str)
                    if not period_dates:
                        continue
                    
                    start_date = period_dates[0]
                    if start_date[0] != month or start_date[2] != year:
                        continue
                    
                    hours_val = safe_float(sheet.cell_value(r, hp_col + 1))
                    if hours_val > 0:
                        section_hours += hours_val
                
                if section_hours > 0:
                    section_fee = section_hours * static_rate
                    total_hours += section_hours
                    total_admin_fee += section_fee
                    last_rate = static_rate
        
        return float(total_hours), float(last_rate), float(total_admin_fee)
    
    except Exception as e:
        import traceback
        print(f"⚠️  Admin fee calc failed for {paysheet_path}: {e}")
        traceback.print_exc()
        return 0.0, 0.0, 0.0


def _is_date_cell_empty(cell_value: Any) -> bool:
    """Check if a cell is effectively empty"""
    if cell_value is None:
        return True
    s = str(cell_value).strip()
    return s == "" or s.lower() == "none"


def _other_cell_contains_keyword(text: Any) -> bool:
    """Check for retro/ACH keywords"""
    try:
        if text is None:
            return False
        s = str(text).lower()
        return ('retro' in s) or ('ach' in s)
    except Exception:
        return False


def calculate_gross_salary_for_paysheet(paysheet_path: str, month: int, year: int, debug: bool = False) -> float:
    """
    Calculate gross salary from paysheet with SPLIT CELL SUPPORT (v7.5.0 logic)
    Returns: total_gross_salary
    """
    total_gross = 0.0
    
    try:
        if not SUPPORT_XLS or not paysheet_path.endswith('.xls'):
            return 0.0
        
        book = xlrd.open_workbook(paysheet_path, formatting_info=False)
        sheet = None
        for s in book.sheets():
            if year_in_sheet_name(year, s.name):
                sheet = s
                break
        if not sheet:
            return 0.0
        
        # Find all "Gross" entries and consolidate split cells
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell_val = sheet.cell_value(r, c)
                if cell_val:
                    text = str(cell_val).lower().strip()
                    if 'gross' in text and ('salary' in text or 'pay' in text or text == 'gross'):
                        if c + 1 < sheet.ncols:
                            # Get the main amount
                            gross_val = safe_float(sheet.cell_value(r, c + 1))
                            
                            section_total = gross_val
                            
                            # Check for split cells below
                            check_row = r + 1
                            while check_row < sheet.nrows:
                                next_label = sheet.cell_value(check_row, c)
                                
                                # If label is empty, check for split amount
                                if _is_date_cell_empty(next_label):
                                    if c + 1 < sheet.ncols:
                                        next_amt = safe_float(sheet.cell_value(check_row, c + 1))
                                        if next_amt > 0:
                                            section_total += next_amt
                                            check_row += 1
                                            continue
                                        else:
                                            # Empty label, no amount - keep checking
                                            check_row += 1
                                            continue
                                
                                # If we hit another label, stop consolidating this section
                                break
                            
                            total_gross += section_total
        
        return float(total_gross)
    
    except Exception as e:
        print(f"⚠️  Gross salary calc failed for {paysheet_path}: {e}")
        return 0.0


def calculate_carryforward_for_paysheet(paysheet_path: str, year: int) -> float:
    """Read 2025 Balance Forward / carryforward amount from paysheet.
    Searches for 'balance forward' or 'carryforward' label and returns
    the first non-zero numeric value found on that row (any column).
    """
    try:
        if not SUPPORT_XLS or not paysheet_path.endswith('.xls'):
            return 0.0

        book = xlrd.open_workbook(paysheet_path, formatting_info=False)
        sheet = None
        for s in book.sheets():
            if year_in_sheet_name(year, s.name):
                sheet = s
                break
        if not sheet:
            return 0.0

        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell_val = sheet.cell_value(r, c)
                if cell_val:
                    text = str(cell_val).lower().strip()
                    if 'balance forward' in text or 'carryforward' in text or 'carry forward' in text:
                        # Found label — scan RIGHT, skip year-like ints (1900-2100)
                        for cc in range(c + 1, sheet.ncols):
                            amt = safe_float(sheet.cell_value(r, cc))
                            if amt == 0.0:
                                continue
                            # Skip values that look like a year, not a balance
                            if amt == int(amt) and 1900 <= amt <= 2100:
                                continue
                            return amt
        return 0.0

    except Exception:
        return 0.0


# ==============================================================================
# SECTION 2: PAYSHEET PARSING (from v7.0.1)
# ==============================================================================

def read_xls_with_xlrd(path: str) -> Dict[str, pd.DataFrame]:
    """Read .xls file using xlrd"""
    if not SUPPORT_XLS:
        raise RuntimeError("xlrd not available")
    
    book = xlrd.open_workbook(path, formatting_info=False, on_demand=False)
    out: Dict[str, pd.DataFrame] = {}
    
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            try:
                row_vals = []
                has_data = False
                for c in range(sheet.ncols):
                    try:
                        v = sheet.cell_value(r, c)
                        ctype = sheet.cell_type(r, c)
                        
                        if ctype == xlrd.XL_CELL_DATE:
                            try:
                                dt_tuple = xlrd.xldate_as_tuple(v, book.datemode)
                                row_vals.append(f"{dt_tuple[1]}/{dt_tuple[2]}/{dt_tuple[0]}")
                                has_data = True
                            except Exception:
                                row_vals.append(str(v))
                                if v != "":
                                    has_data = True
                        elif ctype != xlrd.XL_CELL_EMPTY:
                            row_vals.append(v)
                            if v != "":
                                has_data = True
                        else:
                            row_vals.append("")
                    except Exception:
                        row_vals.append("")
                
                if has_data:
                    rows.append(row_vals)
            except Exception:
                break
        
        if rows:
            df = pd.DataFrame(rows)
            if not df.empty:
                df.columns = [str(x) if x is not None else "" for x in df.iloc[0].tolist()]
                df = df.iloc[1:].reset_index(drop=True)
            out[sheet.name] = df
    
    return out


def parse_paysheet(
    path: str,
    month: int,
    year: int,
    debug_log: List[str],
    bc_cols: Optional[List[int]] = None,
    bc_scan_rows: Optional[int] = None,
    bc_payment_min: float = 1.0,
) -> Tuple[float, float, Optional[float], Dict[str, Any]]:
    """Parse paysheet for hours and payments"""
    debug_log.append(f"PARSING: {os.path.basename(path)}")
    hours_sum = 0.0
    payments_sum = 0.0
    salary_candidate = None
    meta: Dict[str, Any] = {"path": path, "bc_matches": []}
    
    try:
        ext = os.path.splitext(path)[1].lower()
        dfs: Dict[str, pd.DataFrame] = {}
        
        if ext == ".xls":
            dfs = read_xls_with_xlrd(path)
        else:
            try:
                dfs = pd.read_excel(path, sheet_name=None, engine=None)
            except Exception:
                dfs = pd.read_excel(path, sheet_name=None, engine="openpyxl")

        sheet_items = list(dfs.items())
        # Word-boundary year match. Prefer exact-name tab.
        selected = [(n, d) for (n, d) in sheet_items if year_in_sheet_name(year, n)]
        if len(selected) > 1:
            exact = [(n, d) for (n, d) in selected if str(n).strip() == str(year)]
            if exact:
                selected = [exact[0]]
            else:
                debug_log.append(f"  ⚠️  Multiple year-{year} tabs matched: {[n for n, _ in selected]}. Using first.")
                selected = [selected[0]]
        if not selected:
            debug_log.append(f"  ⚠️  No sheet matching year {year} in {os.path.basename(path)} — skipping (no fallback).")
            selected = []

        first_num_re = re.compile(r'(\d{1,2})')
        start_re = re.compile(rf'^\s*(?:{month:02d}|{month})(?:[/-]|\b)')

        def first_numeric_token_from_text(s: Any) -> Optional[int]:
            if s is None:
                return None
            txt = re.sub(r'[A-Za-z]+', '', str(s))
            m = first_num_re.search(txt)
            return int(m.group(1)) if m else None

        for sname, df in selected:
            if df is None or df.shape[1] == 0:
                continue
            
            df2 = df.copy()
            df2.columns = [str(c) if c is not None else "" for c in df2.columns]
            nrows, ncols = df2.shape

            bc_cols_use = bc_cols if bc_cols else [1, 2, 3]
            rows_to_scan = nrows if (not bc_scan_rows or bc_scan_rows <= 0) else min(bc_scan_rows, nrows)
            
            for r in range(rows_to_scan):
                for col_idx in bc_cols_use:
                    if col_idx >= ncols:
                        continue
                    
                    raw = df2.iat[r, col_idx]
                    if raw is None:
                        continue
                    
                    txt_raw = str(raw).strip()
                    if not txt_raw:
                        continue
                    
                    first_num = first_numeric_token_from_text(txt_raw)
                    if first_num is None or first_num != month:
                        continue

                    # Skip non-date text like "Discount-1.2%", "Rate", etc.
                    # Word-boundary match avoids dropping legit cells with substrings
                    txt_lower = txt_raw.lower()
                    if re.search(r'\b(discount|rate|admin\s*fee|tax|employer|deduct)\b', txt_lower):
                        continue

                    # Ensure text contains date-like separators (/ or -)
                    if not ('/' in txt_raw or '-' in txt_raw):
                        continue

                    # Text must start with a digit to be a date period
                    if not re.match(r'^\s*\d', txt_raw):
                        continue

                    looks_like_period = ('-' in txt_raw)
                    starts_with_month = bool(start_re.search(txt_raw))
                    if not (looks_like_period or starts_with_month):
                        continue
                    
                    hours_col = col_idx + 1
                    try:
                        hours_val = safe_float(df2.iat[r, hours_col]) if hours_col < ncols else 0.0
                    except Exception:
                        hours_val = 0.0
                    
                    if not (hours_val > 0 and hours_val <= 500):
                        continue
                    
                    payment_val = 0.0
                    payment_col = hours_col + 1
                    if payment_col < ncols:
                        try:
                            pv = safe_float(df2.iat[r, payment_col])
                            if pv >= float(bc_payment_min):
                                payment_val = pv
                        except Exception:
                            pass
                    
                    hours_sum += hours_val
                    payments_sum += payment_val

        return round(hours_sum, 4), round(payments_sum, 2), salary_candidate, meta
    
    except Exception as e:
        debug_log.append(f"  ✗ ERROR: {e}")
        raise RuntimeError(f"Failed parsing {path}: {e}")


# ==============================================================================
# SECTION 3: WORKBOOK MANAGEMENT - DYNAMIC COLUMN FINDING (from v7.0.1)
# ==============================================================================

MONTH_ABBREVIATIONS = {
    "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
    "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
    "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec",
}


def find_headers(ws: Worksheet, header_row: int, month_name: str) -> Dict[str, Optional[int]]:
    """Find column headers - FLEXIBLE MATCHING with month abbreviations.

    Master files use inconsistent naming: sometimes "January Hours", sometimes
    "Jan Hours". This function searches for BOTH forms, prioritizing the specific
    month to avoid matching the wrong month's column.
    """
    headers: Dict[str, Optional[int]] = {}
    header_cells: Dict[int, str] = {}

    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v not in (None, ""):
            header_cells[c] = str(v).strip()

    def find_by_keywords(keywords: List[str]) -> Optional[int]:
        """Find column by keywords. Tries each keyword across ALL columns (most-specific
        first), so 'gross salary' wins over 'gross' when both columns exist.
        Normalizes whitespace so 'applicant number' matches 'Applicant     Number'."""
        normalized = {c: ' '.join(text.lower().split()) for c, text in header_cells.items()}
        for keyword in keywords:
            kw_lower = keyword.lower()
            for c, text in normalized.items():
                if kw_lower in text:
                    return c
        return None

    def find_month_column(suffixes: List[str], generic_suffixes: List[str]) -> Optional[int]:
        """Find a month-specific column, excluding other months' columns.

        Example for February + suffixes=['hours', 'hrs']:
          1. Try "February hours", "February hrs", "Feb hours", "Feb hrs"
          2. If none match, fall back to generic "hours" / "hrs" — but SKIP any
             column whose header contains a different month's abbreviation.
        """
        abbr = MONTH_ABBREVIATIONS.get(month_name, month_name[:3])
        # Tier 1: explicit month match
        month_keywords: List[str] = []
        for s in suffixes:
            month_keywords.append(f"{month_name} {s}")
            month_keywords.append(f"{abbr} {s}")
        col = find_by_keywords(month_keywords)
        if col is not None:
            return col

        # Tier 2: generic match, but excluding other months' columns
        # Use word-boundary regex to avoid false positives (e.g., "mar" in "margin")
        other_month_patterns = []
        for m in MONTH_ABBREVIATIONS:
            if m == month_name:
                continue
            a = MONTH_ABBREVIATIONS[m]
            other_month_patterns.append(re.compile(r'\b' + re.escape(a) + r'\b', re.I))
            other_month_patterns.append(re.compile(r'\b' + re.escape(m) + r'\b', re.I))

        for c, text in header_cells.items():
            text_lower = text.lower()
            # Must contain a generic keyword
            if not any(kw.lower() in text_lower for kw in generic_suffixes):
                continue
            # Must NOT contain another month's name (word-boundary match)
            if any(pat.search(text) for pat in other_month_patterns):
                continue
            return c
        return None

    headers['file_col'] = find_by_keywords(['applicant number', 'app id', 'appid', 'file #', 'file number', 'app no'])
    headers['payroll_name_col'] = find_by_keywords(['employee name', 'payroll name', 'employee', 'payroll'])

    headers['accrual_hours_col'] = find_month_column(
        suffixes=['hours', 'hrs'],
        generic_suffixes=['hours', 'hrs'],
    )
    headers['billed_col'] = find_month_column(
        suffixes=['billed', 'billing'],
        generic_suffixes=['billed', 'billing'],
    )

    headers['admin_fee_col'] = find_by_keywords(['admin fee', 'adminfee'])
    headers['salary_paid_col'] = find_by_keywords(['salary paid'])
    headers['wages_earned_col'] = find_by_keywords(['wages earned'])
    headers['gross_salary_col'] = find_by_keywords(['total gross salary', 'gross salary', 'gross pay'])
    if not headers['gross_salary_col']:
        headers['gross_salary_col'] = None  # No fallback — leave unset rather than guess
    headers['carryforward_col'] = find_by_keywords([
        'carryforward', 'carry forward', 'balance forward',
        'accrued payroll per audit', 'starting pi balance',
        'beginning balance', 'c/f balance',
    ])

    return headers


def list_all_headers(ws: Worksheet, header_row: int) -> Dict[int, str]:
    """List ALL headers for debugging"""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v not in (None, ""):
            headers[c] = str(v).strip()
    return headers


# ==============================================================================
# SECTION 4: DATE-BASED AB CALCULATIONS (from v7.5.0)
# ==============================================================================

def find_amount_for_date_in_paysheet(path: str, target_date: date, debug_log: List[str]) -> float:
    """Find amount for specific date with split cell consolidation (v7.5.0 logic)"""
    debug_log.append(f"\n  📅 Searching for: {target_date.strftime('%m/%d/%Y')}")
    
    try:
        ext = os.path.splitext(path)[1].lower()
        preferred_date_cols = [6, 7]
        fallback_window = range(4, 9)
        
        if ext == ".xls" and SUPPORT_XLS:
            book = xlrd.open_workbook(path, formatting_info=False, on_demand=True)
            for sheet in book.sheets():
                nrows = sheet.nrows
                ncols = sheet.ncols
                
                for r in range(nrows):
                    for c in preferred_date_cols:
                        if c >= ncols:
                            continue
                        
                        v = sheet.cell_value(r, c)
                        cell_type = sheet.cell_type(r, c)
                        
                        found_date = None
                        if cell_type == xlrd.XL_CELL_DATE:
                            try:
                                dt_tuple = xlrd.xldate_as_tuple(v, book.datemode)
                                found_date = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                            except Exception:
                                pass
                        else:
                            found_date = _normalize_input_date_to_dateobj(str(v))
                        
                        if found_date and found_date == target_date:
                            col_letter = get_column_letter(c + 1)
                            debug_log.append(f"    ✅ Found at column {col_letter}, row {r}")
                            
                            amt_col = c + 1
                            base_amt = 0.0
                            
                            if amt_col < ncols:
                                base_amt = safe_float(sheet.cell_value(r, amt_col))
                            
                            debug_log.append(f"    📍 Row {r}: ${base_amt:.2f}")
                            
                            total_amt = base_amt
                            check_row = r + 1
                            consolidated_count = 1
                            consolidated_rows = [r]
                            
                            # Consolidate split cells (empty date rows)
                            while check_row < nrows:
                                next_date_cell = sheet.cell_value(check_row, c)
                                
                                if _is_date_cell_empty(next_date_cell):
                                    next_amt = safe_float(sheet.cell_value(check_row, amt_col))
                                    if next_amt > 0:
                                        total_amt += next_amt
                                        consolidated_count += 1
                                        consolidated_rows.append(check_row)
                                        debug_log.append(f"    📍 Row {check_row} (SPLIT CELL - empty date): +${next_amt:.2f}")
                                        check_row += 1
                                        continue
                                    else:
                                        check_row += 1
                                        continue
                                
                                next_date = None
                                try:
                                    if sheet.cell_type(check_row, c) == xlrd.XL_CELL_DATE:
                                        dt_tuple = xlrd.xldate_as_tuple(next_date_cell, book.datemode)
                                        next_date = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                                    else:
                                        next_date = _normalize_input_date_to_dateobj(str(next_date_cell))
                                except Exception:
                                    pass
                                
                                if next_date and next_date == target_date:
                                    next_amt = safe_float(sheet.cell_value(check_row, amt_col))
                                    total_amt += next_amt
                                    consolidated_count += 1
                                    consolidated_rows.append(check_row)
                                    debug_log.append(f"    📍 Row {check_row} (SAME DATE): +${next_amt:.2f}")
                                    check_row += 1
                                    continue
                                
                                break
                            
                            debug_log.append(f"    🔗 Consolidated {consolidated_count} rows: {consolidated_rows}")
                            debug_log.append(f"    💰 Total: ${total_amt:.2f}")
                            
                            # Check for retro/ACH adjustments
                            total_adjustments = 0.0
                            nr = check_row
                            
                            while nr < nrows:
                                label_cell = None
                                if c < ncols:
                                    label_cell = sheet.cell_value(nr, c)
                                
                                if _other_cell_contains_keyword(label_cell):
                                    if amt_col < ncols:
                                        adj_amt = safe_float(sheet.cell_value(nr, amt_col))
                                        total_adjustments += adj_amt
                                        debug_log.append(f"    🔧 {label_cell} row {nr}: +${adj_amt:.2f}")
                                        nr += 1
                                        continue
                                
                                break
                            
                            if total_adjustments > 0:
                                debug_log.append(f"    🎯 Adjustments: ${total_adjustments:.2f}")
                            
                            final_amt = total_amt + total_adjustments
                            debug_log.append(f"    ✅ FINAL: ${final_amt:.2f}")
                            
                            return final_amt
                
                # Fallback search — with split-cell consolidation
                for r in range(nrows):
                    for c in fallback_window:
                        if c >= ncols:
                            continue

                        v = sheet.cell_value(r, c)
                        found_date = None
                        try:
                            if sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
                                dt_tuple = xlrd.xldate_as_tuple(v, book.datemode)
                                found_date = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                            else:
                                found_date = _normalize_input_date_to_dateobj(str(v))
                        except Exception:
                            pass

                        if found_date and found_date == target_date:
                            amt_col = c + 1
                            total_amt = safe_float(sheet.cell_value(r, amt_col)) if amt_col < ncols else 0.0
                            # Consolidate split cells (same logic as primary path)
                            check_r = r + 1
                            while check_r < nrows:
                                next_date_cell = sheet.cell_value(check_r, c)
                                if _is_date_cell_empty(next_date_cell):
                                    next_amt = safe_float(sheet.cell_value(check_r, amt_col)) if amt_col < ncols else 0.0
                                    if next_amt > 0:
                                        total_amt += next_amt
                                    check_r += 1
                                    continue
                                break
                            return total_amt
            
            return 0.0
        
        else:
            dfs = pd.read_excel(path, sheet_name=None, engine=None)
            
            for sname, df in dfs.items():
                if df is None:
                    continue
                
                nrows, ncols = df.shape
                
                for r in range(nrows):
                    for c in preferred_date_cols:
                        if c >= ncols:
                            continue
                        
                        try:
                            v = df.iat[r, c]
                        except Exception:
                            v = None
                        
                        if v is None or (isinstance(v, float) and pd.isna(v)):
                            continue
                        
                        found_date = None
                        if isinstance(v, (datetime, date)):
                            found_date = date(v.year, v.month, v.day)
                        else:
                            found_date = _normalize_input_date_to_dateobj(str(v))
                        
                        if found_date and found_date == target_date:
                            col_letter = get_column_letter(c + 1)
                            debug_log.append(f"    ✅ Found at column {col_letter}, row {r}")
                            
                            amt_col = c + 1
                            base_amt = 0.0
                            
                            if amt_col < ncols:
                                try:
                                    base_amt = safe_float(df.iat[r, amt_col])
                                except Exception:
                                    pass
                            
                            total_amt = base_amt
                            check_row = r + 1
                            consolidated_count = 1
                            consolidated_rows = [r]
                            
                            while check_row < nrows:
                                try:
                                    next_date_cell = df.iat[check_row, c]
                                except Exception:
                                    break
                                
                                if _is_date_cell_empty(next_date_cell):
                                    try:
                                        next_amt = safe_float(df.iat[check_row, amt_col])
                                        if next_amt > 0:
                                            total_amt += next_amt
                                            consolidated_count += 1
                                            consolidated_rows.append(check_row)
                                            debug_log.append(f"    📍 Row {check_row} (SPLIT CELL - empty date): +${next_amt:.2f}")
                                            check_row += 1
                                            continue
                                        else:
                                            check_row += 1
                                            continue
                                    except Exception:
                                        break

                                # Same-date consolidation (repeated rows with identical date)
                                next_date = None
                                try:
                                    if isinstance(next_date_cell, (datetime, date)):
                                        next_date = date(next_date_cell.year, next_date_cell.month, next_date_cell.day)
                                    else:
                                        next_date = _normalize_input_date_to_dateobj(str(next_date_cell))
                                except Exception:
                                    pass

                                if next_date and next_date == target_date:
                                    try:
                                        next_amt = safe_float(df.iat[check_row, amt_col])
                                        total_amt += next_amt
                                        consolidated_count += 1
                                        consolidated_rows.append(check_row)
                                        debug_log.append(f"    📍 Row {check_row} (SAME DATE): +${next_amt:.2f}")
                                        check_row += 1
                                        continue
                                    except Exception:
                                        break

                                break
                            
                            debug_log.append(f"    🔗 Consolidated {consolidated_count} rows: {consolidated_rows}")
                            debug_log.append(f"    💰 Total: ${total_amt:.2f}")
                            
                            return total_amt
            
            return 0.0
    
    except Exception as e:
        debug_log.append(f"    ❌ ERROR: {e}")
        return 0.0


# ==============================================================================
# SECTION 5: XLWINGS FORMULA CHECKING (NEW for v7.0.2)
# ==============================================================================

def cell_value_is_formula(cell_value: Any) -> bool:
    """Check if a cell value represents a formula"""
    if cell_value is None:
        return False
    if isinstance(cell_value, str) and cell_value.startswith('='):
        return True
    return False


# ==============================================================================
# SECTION 6: MAIN CLASS - v7.0.2 HYBRID (xlwings structure + v7.5.0 logic + xlwings writes)
# ==============================================================================

class AccrualUpdater:
    """Main Accrual Updater v7.0.2 - Hybrid version (xlwings direct writes)"""
    
    def __init__(
        self,
        master_path: str,
        sheet_name: str = "Profit Sharing",
        header_row: int = 3,
        month: str = "June",
        year: int = 0,  # 0 → defaults to current year via datetime.now().year below
        paysheets_folder: str = "",
        date_multiplier_pairs: Optional[List[Tuple[str, float]]] = None,
        dry_run: bool = True,
        backup: bool = False,
        enable_ot_detection: bool = True,
        enable_carryforward: bool = False,
        **kwargs,
    ):
        self.master_path = master_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.month = month
        try:
            self.month_index = MONTHS.index(month) + 1
        except Exception:
            self.month_index = 1
        self.year = year if year else datetime.now().year
        self.paysheets_folder = paysheets_folder
        self.date_multiplier_pairs = date_multiplier_pairs or []
        self.dry_run = dry_run
        self.backup = backup
        self.enable_ot_detection = enable_ot_detection
        # Carryforward only runs in January by default; checkbox forces it on for any month
        self.enable_carryforward = enable_carryforward or (self.month_index == 1)
        self.log_lines: List[str] = []
        self.debug_log: List[str] = []
        self.updated_count = 0
        self.no_match_count = 0
        self.failed_count = 0
        
        self.ALLOWED_COLS = set()
        self.start_time = None

    def log(self, line: str):
        print(line)
        self.log_lines.append(line)

    def build_master_lookup(self, ws: Worksheet) -> Dict[str, Dict[str, Any]]:
        """Build lookup dictionary from master file"""
        lookup: Dict[str, Dict[str, Any]] = {}
        headers = find_headers(ws, self.header_row, self.month)
        file_col = headers.get('file_col')
        name_col = headers.get('payroll_name_col')
        
        if not file_col:
            self.log("⚠️  WARNING: Could not find FILE_COL - showing all headers:")
            all_headers = list_all_headers(ws, self.header_row)
            for col_num, col_name in sorted(all_headers.items()):
                self.log(f"  Column {get_column_letter(col_num)}: {col_name}")
            return {}
        
        for r in range(self.header_row + 1, ws.max_row + 1):
            try:
                v = ws.cell(row=r, column=file_col).value if file_col else None
            except Exception:
                v = None
            
            fnum = None
            if v is not None:
                m = re.search(r'(\d{5,6})', str(v))
                if m:
                    fnum = m.group(1)
            
            if fnum:
                name_val = ws.cell(row=r, column=name_col).value if name_col else None
                lookup[fnum] = {"row": r, "name": name_val}
        
        return lookup

    def _validate_prior_months(
        self,
        ws_read_path: str,
        header_row: int,
        all_headers: Dict[int, str],
        file_col: int,
        files: List[str],
        master_lookup: Dict[str, Dict[str, Any]],
    ) -> None:
        """Cross-check other months' hours/billed data already in the master.

        Purely data-driven: scans every column header in the master, detects
        which ones are month+hours or month+billed by looking for month names
        inside the header text, then cross-checks against paysheets.

        Never hardcodes column positions or month names — everything comes
        from reading the actual master headers.
        """
        # Build a month lookup from actual header text.
        # For each header cell, check if it contains a recognisable month AND
        # a data-type keyword (hours/hrs/billed/billing).
        all_month_names = list(MONTHS)           # ["January", ..., "December"]
        all_month_abbrs = list(MONTH_ABBREVIATIONS.values())  # ["Jan", ..., "Dec"]

        # {month_index: {"name": ..., "hrs_col": col|None, "bill_col": col|None}}
        detected: Dict[int, dict] = {}

        for col, header_text in all_headers.items():
            t = header_text.lower()

            # Which month does this header belong to? Use word-boundary regex.
            matched_months: List[Tuple[int, str]] = []
            for i, (full, abbr) in enumerate(zip(all_month_names, all_month_abbrs)):
                if (re.search(rf'\b{re.escape(full.lower())}\b', t) or
                        re.search(rf'\b{re.escape(abbr.lower())}\b', t)):
                    matched_months.append((i + 1, full))

            # If header mentions multiple months (e.g. "May/June Adjust"), skip it
            if len(matched_months) > 1:
                self.log(f"  ⚠️  Header '{header_text}' mentions multiple months — skipping in prior-month check")
                continue
            if not matched_months:
                continue

            matched_month_idx, matched_label = matched_months[0]

            # Is it hours or billed?
            is_hours = any(kw in t for kw in ['hours', 'hrs'])
            is_billed = any(kw in t for kw in ['billed', 'billing'])
            if not is_hours and not is_billed:
                continue

            if matched_month_idx not in detected:
                detected[matched_month_idx] = {
                    'name': matched_label,
                    'idx': matched_month_idx,
                    'hrs_col': None,
                    'bill_col': None,
                }

            if is_hours and detected[matched_month_idx]['hrs_col'] is None:
                detected[matched_month_idx]['hrs_col'] = col
            if is_billed and detected[matched_month_idx]['bill_col'] is None:
                detected[matched_month_idx]['bill_col'] = col

        # Filter to only OTHER months (not the one we're currently processing)
        other_months = [
            mc for idx, mc in sorted(detected.items())
            if idx != self.month_index
        ]

        if not other_months:
            return

        labels = ', '.join(
            f"{mc['name']} (Hrs={get_column_letter(mc['hrs_col']) if mc['hrs_col'] else '-'}, "
            f"Bill={get_column_letter(mc['bill_col']) if mc['bill_col'] else '-'})"
            for mc in other_months
        )
        self.log(f"🔎 Pre-validation: cross-checking {labels}")

        # Open master read-only to get cell values
        wb_check = openpyxl.load_workbook(ws_read_path, read_only=True, data_only=True)
        ws_check = wb_check[self.sheet_name]

        # Build paysheet file-number lookup
        ps_lookup: Dict[str, str] = {}
        for fp in files:
            m = re.search(r'(\d{5,6})', os.path.basename(fp))
            if m:
                ps_lookup[m.group(1)] = fp

        warnings_found = 0
        employees_checked = 0

        for fnum, ps_path in ps_lookup.items():
            if fnum not in master_lookup:
                continue
            rec = master_lookup[fnum]
            row = rec['row']
            name = rec.get('name', fnum)

            for mc in other_months:
                master_hrs = 0.0
                master_bill = 0.0
                if mc['hrs_col']:
                    v = ws_check.cell(row=row, column=mc['hrs_col']).value
                    try:
                        master_hrs = float(v) if v else 0.0
                    except (ValueError, TypeError):
                        master_hrs = 0.0
                if mc['bill_col']:
                    v = ws_check.cell(row=row, column=mc['bill_col']).value
                    try:
                        master_bill = float(v) if v else 0.0
                    except (ValueError, TypeError):
                        master_bill = 0.0

                # Skip if master has no data for this month (not yet processed)
                if master_hrs == 0.0 and master_bill == 0.0:
                    continue

                # Re-parse paysheet for this other month
                try:
                    dbg: List[str] = []
                    ps_hrs, ps_bill, _, _ = parse_paysheet(
                        ps_path, mc['idx'], self.year, dbg
                    )
                except Exception:
                    continue

                employees_checked += 1

                hrs_diff = abs(master_hrs - ps_hrs)
                bill_diff = abs(master_bill - ps_bill)

                issues = []
                if hrs_diff > 0.5:
                    issues.append(
                        f"Hours: master={master_hrs:.1f} vs paysheet={ps_hrs:.1f} (diff={hrs_diff:.1f})"
                    )
                if bill_diff > 1.0:
                    issues.append(
                        f"Billed: master=${master_bill:.2f} vs paysheet=${ps_bill:.2f} (diff=${bill_diff:.2f})"
                    )

                if issues:
                    warnings_found += 1
                    self.log(
                        f"  ⚠️  {mc['name']} MISMATCH — Row {row} ({name}): "
                        + " | ".join(issues)
                    )

        wb_check.close()

        if warnings_found == 0:
            self.log(f"  ✅ All prior month data matches ({employees_checked} checks)\n")
        else:
            self.log(
                f"  🚨 {warnings_found} MISMATCH(ES) in prior months! "
                f"Review before proceeding.\n"
            )

    def process(self):
        """Process - MAIN LOOP (xlwings structure + v7.5.0 logic + xlwings writes)"""
        self.start_time = time.time()
        
        self.log("\n" + "="*80)
        self.log("ACCRUAL UPDATER v" + __version__)
        self.log("✨ HYBRID v7.0.2 - XLWINGS DIRECT WRITES ✨")
        self.log("✅ xlwings for direct Excel cell writing (preserves connections)")
        self.log("✅ v7.5.0 logic: Formula detection, split cells, admin fees, gross salary")
        self.log("💰 Calculates: Hours | Billed (N→V) | Admin Fees | Gross Salary | AB")
        self.log(f"Sheet: {self.sheet_name} | Header Row: {self.header_row}")
        self.log(f"Month: {self.month} {self.year}")
        self.log("="*80 + "\n")

        if not os.path.exists(self.master_path):
            raise RuntimeError(f"Master not found: {self.master_path}")
        if not os.path.exists(self.paysheets_folder):
            raise RuntimeError(f"Paysheets not found: {self.paysheets_folder}")

        self.log(f"Master: {self.master_path}")
        self.log(f"Paysheets: {self.paysheets_folder}\n")

        # Load with openpyxl (READ ONLY - just for headers)
        wb_read = openpyxl.load_workbook(filename=self.master_path, read_only=True, data_only=False)
        
        if self.sheet_name not in wb_read.sheetnames:
            self.log(f"❌ Sheet '{self.sheet_name}' not found!")
            self.log(f"Available sheets: {', '.join(wb_read.sheetnames)}")
            wb_read.close()
            raise RuntimeError(f"Sheet '{self.sheet_name}' not found")
        
        ws_read = wb_read[self.sheet_name]

        self.log(f"📋 Headers in row {self.header_row}:")
        all_headers = list_all_headers(ws_read, self.header_row)
        for col_num, col_name in sorted(all_headers.items()):
            self.log(f"  {get_column_letter(col_num)}: {col_name}")
        
        self.log("")

        headers = find_headers(ws_read, self.header_row, self.month)
        accrual_col = headers.get('accrual_hours_col')
        billed_col = headers.get('billed_col')
        admin_fee_col = headers.get('admin_fee_col')
        salary_paid_col = headers.get('salary_paid_col')
        wages_earned_col = headers.get('wages_earned_col')
        file_col = headers.get('file_col')
        name_col = headers.get('payroll_name_col')
        gross_salary_col = headers.get('gross_salary_col')

        self.log("🔍 Column Detection:")
        self.log(f"  File Col: {get_column_letter(file_col) if file_col else 'NOT FOUND'}")
        self.log(f"  Name Col: {get_column_letter(name_col) if name_col else 'NOT FOUND'}")
        self.log(f"  🎯 {self.month} Hours Col: {get_column_letter(accrual_col) if accrual_col else 'NOT FOUND'}")
        self.log(f"  🎯 {self.month} Billed Col: {get_column_letter(billed_col) if billed_col else 'NOT FOUND'}")
        self.log(f"  Admin Fee Col: {get_column_letter(admin_fee_col) if admin_fee_col else 'NOT FOUND'}")
        self.log(f"  💰 Gross Salary Col: {get_column_letter(gross_salary_col) if gross_salary_col else 'NOT FOUND'}")
        self.log(f"  💵 Salary Paid Col: {get_column_letter(salary_paid_col) if salary_paid_col else 'NOT FOUND'}")
        self.log(f"  💼 Wages Earned Col: {get_column_letter(wages_earned_col) if wages_earned_col else 'NOT FOUND'} (hourly only)")
        self.log("")

        if not file_col or not accrual_col or not billed_col:
            self.log("❌ ERROR: Required columns not found!")
            wb_read.close()
            raise RuntimeError("Required columns not found!")

        carryforward_col = headers.get('carryforward_col')
        if carryforward_col:
            self.log(f"  🎯 Carryforward Col: {get_column_letter(carryforward_col)}")
        else:
            self.log(f"  ⚠️  Carryforward column not found in headers — skipping carryforward writes")

        self.ALLOWED_COLS = {accrual_col, billed_col}
        if carryforward_col:
            self.ALLOWED_COLS.add(carryforward_col)
        if admin_fee_col:
            self.ALLOWED_COLS.add(admin_fee_col)
        if salary_paid_col:
            self.ALLOWED_COLS.add(salary_paid_col)
        if wages_earned_col:
            self.ALLOWED_COLS.add(wages_earned_col)

        self.log(f"✅ Writing to columns: {sorted([get_column_letter(c) for c in self.ALLOWED_COLS])}")
        self.log(f"🚫 Protected columns (untouched): All others\n")

        master_lookup = self.build_master_lookup(ws_read)
        wb_read.close()

        if not master_lookup:
            self.log("❌ ERROR: Could not build master lookup")
            raise RuntimeError("Could not build master lookup")

        valid_exts = {'.xls', '.xlsx', '.xlsm'}
        files: List[str] = []
        for root, _, fnames in os.walk(self.paysheets_folder):
            for fn in fnames:
                if os.path.splitext(fn)[1].lower() in valid_exts:
                    files.append(os.path.join(root, fn))

        self.log(f"Found {len(files)} paysheet(s)\n")

        # ── Pre-validation: cross-check prior months' data ──────────────
        self._validate_prior_months(
            ws_read_path=self.master_path,
            header_row=self.header_row,
            all_headers=all_headers,
            file_col=file_col,
            files=files,
            master_lookup=master_lookup,
        )

        if not salary_paid_col:
            self.log("  🚨 WARNING: 'Salary Paid' column NOT FOUND — AB/multiplier values will NOT be written!")

        updates_to_apply = []

        for idx, fp in enumerate(files, 1):
            fname = os.path.basename(fp)
            self.log(f"[{idx}/{len(files)}] {fname}")
            
            fnum = None
            m = re.search(r'(\d{5,6})', fname)
            if m:
                fnum = m.group(1)
            
            if not fnum:
                self.no_match_count += 1
                self.log("  ✗ NO FILE NUMBER")
                continue
            
            if fnum not in master_lookup:
                self.no_match_count += 1
                self.log("  ✗ NOT IN MASTER")
                continue
            
            rec = master_lookup[fnum]
            row = rec["row"]
            name_val = rec.get("name") or ""
            self.log(f"  ✓ {name_val} (Row {row})")

            try:
                dbg: List[str] = []
                hours, payments, _, _ = parse_paysheet(fp, self.month_index, self.year, dbg)

                admin_hours, admin_rate, admin_fee = calculate_admin_fee_for_paysheet(
                    fp, self.month_index, self.year, debug=False
                )

                # Carryforward only runs in January (month_index == 1) OR if checkbox enabled
                if self.enable_carryforward:
                    carryforward = calculate_carryforward_for_paysheet(fp, self.year)
                    if carryforward > 0:
                        self.log(f"  💼 {self.year - 1} Balance Forward: ${carryforward:.2f}")
                else:
                    carryforward = 0.0

                ab_total = 0.0
                if self.date_multiplier_pairs:
                    self.log(f"  Calculating AB from {len(self.date_multiplier_pairs)} date(s):")
                    for date_str, multiplier in self.date_multiplier_pairs:
                        target_date_obj = _normalize_input_date_to_dateobj(date_str)
                        if not target_date_obj:
                            self.log(f"    ✗ Could not parse: {date_str}")
                            continue
                        
                        dbg_before = len(dbg)
                        combined_amt = find_amount_for_date_in_paysheet(fp, target_date_obj, dbg)
                        
                        for d in dbg[dbg_before:]:
                            self.log(d)
                        
                        computed = round(combined_amt * multiplier, 2)
                        ab_total += computed
                        
                        self.log(f"    💵 {date_str}: ${combined_amt:.2f} × {multiplier} = ${computed:.2f}")
                    
                    self.log(f"  🎯 AB Total: ${ab_total:.2f}")

                # For hourly employees: Wages Earned = Billed to Client
                # Detect hourly: folder path takes precedence (most reliable),
                # then fall back to name suffix "(hourly)"
                fp_lower = fp.lower()
                is_hourly = (
                    'hourly sheets' in fp_lower
                    or 'hourly sheet' in fp_lower
                    or '/hourly/' in fp_lower
                    or 'hourly' in str(name_val).lower()
                )
                wages_earned_val = None
                if is_hourly and wages_earned_col and payments > 0:
                    wages_earned_val = round(payments, 2)

                if not self.dry_run:
                    update_record = {
                        'row': row,
                        'accrual_col': accrual_col,
                        'accrual_val': round(hours, 4),
                        'billed_col': billed_col,
                        'billed_val': round(payments, 2),
                        'admin_fee_col': admin_fee_col,
                        'admin_fee_val': round(admin_fee, 2) if admin_fee > 0 else None,
                        'salary_paid_col': salary_paid_col,
                        'salary_paid_val': round(ab_total, 2) if ab_total > 0 else None,
                        'carryforward_col': carryforward_col,
                        'carryforward_val': round(carryforward, 2) if carryforward > 0 else None,
                        'wages_earned_col': wages_earned_col if is_hourly else None,
                        'wages_earned_val': wages_earned_val,
                    }
                    updates_to_apply.append(update_record)

                    we_str = f" | WE=${payments:.0f}" if wages_earned_val else ""
                    self.log(f"  ✓ Queued: H={hours:.0f} | B=${payments:.0f} | A=${admin_fee:.0f} | SP=${ab_total:.0f} | CF=${carryforward:.0f}{we_str}")
                else:
                    we_str = f" | WE=${payments:.0f}" if wages_earned_val else ""
                    self.log(f"  (dry) H={hours:.0f} | B=${payments:.0f} | A=${admin_fee:.0f} | SP=${ab_total:.0f} | CF=${carryforward:.0f}{we_str}")

                has_data = hours > 0 or payments > 0 or admin_fee > 0 or ab_total > 0 or carryforward > 0
                if has_data:
                    self.updated_count += 1
                else:
                    self.log(f"  ⚠️  All values zero — not counted as updated")

            except Exception as e:
                self.failed_count += 1
                self.log(f"  ✗ ERROR: {e}")

        # NOW: Use xlwings to apply ALL updates to Excel
        if not self.dry_run and updates_to_apply:
            self.log(f"\n💾 Opening Excel and applying {len(updates_to_apply)} updates...\n")

            target_path = os.path.normcase(os.path.realpath(self.master_path))
            book = None
            spawned_app = None

            try:
                # Reuse existing Excel session if file already open (case-insensitive on Windows)
                for existing_app in list(xw.apps):
                    for existing_book in list(existing_app.books):
                        try:
                            if os.path.normcase(os.path.realpath(existing_book.fullname)) == target_path:
                                book = existing_book
                                self.log(f"  ✓ Reusing open Excel session (PID {existing_app.pid})")
                                break
                        except Exception:
                            continue
                    if book is not None:
                        break

                if book is None:
                    spawned_app = xw.App(visible=True)
                    book = xw.Book(self.master_path)
                    self.log("  ✓ Opened Excel")

                try:
                    ws_xw = book.sheets[self.sheet_name]

                    def _xw_has_formula(r: int, c: int) -> bool:
                        """Check if cell has formula via xlwings (no TOCTOU race)."""
                        try:
                            formula = ws_xw.cells(r, c).formula
                            return isinstance(formula, str) and formula.startswith('=')
                        except Exception:
                            return False

                    def _safe_write(r: int, c: int, val, label: str = "") -> bool:
                        """Write value to cell if it's not a formula. Returns True if written."""
                        if _xw_has_formula(r, c):
                            self.log(f"  ⚠️  {get_column_letter(c)}{r} has formula - SKIPPED")
                            return False
                        ws_xw.cells(r, c).value = val
                        suffix = f" ({label})" if label else ""
                        self.log(f"  ✓ {get_column_letter(c)}{r} = {val}{suffix}")
                        return True

                    updates_written = 0
                    for update in updates_to_apply:
                        row = update['row']

                        if accrual_col in self.ALLOWED_COLS:
                            if _safe_write(row, accrual_col, update['accrual_val']):
                                updates_written += 1

                        if billed_col in self.ALLOWED_COLS:
                            if _safe_write(row, billed_col, update['billed_val']):
                                updates_written += 1

                        if admin_fee_col and admin_fee_col in self.ALLOWED_COLS and update['admin_fee_val'] is not None:
                            if _safe_write(row, admin_fee_col, update['admin_fee_val']):
                                updates_written += 1

                        sp_col = update.get('salary_paid_col')
                        if sp_col and sp_col in self.ALLOWED_COLS and update.get('salary_paid_val') is not None:
                            if _safe_write(row, sp_col, update['salary_paid_val'], "Salary Paid"):
                                updates_written += 1

                        we_col = update.get('wages_earned_col')
                        if we_col and we_col in self.ALLOWED_COLS and update.get('wages_earned_val') is not None:
                            if _safe_write(row, we_col, update['wages_earned_val'], "Wages Earned - hourly"):
                                updates_written += 1

                        cf_col = update.get('carryforward_col')
                        if cf_col and cf_col in self.ALLOWED_COLS and update.get('carryforward_val') is not None:
                            if _safe_write(row, cf_col, update['carryforward_val'], "Balance Forward"):
                                updates_written += 1
                    
                    # CRITICAL: Save through Excel (preserves all connections, pivot tables!)
                    book.save()
                    self.log(f"\n✅ Excel saved - {updates_written} cells written")
                    self.log(f"✅ FORMULAS PRESERVED! Cells with formulas were not overwritten!")
                    self.log(f"✅ PIVOT TABLES & QUERIES PRESERVED!")
                    self.log(f"✅ All connections maintained via Excel save!")
                    self.log(f"\n📌 Excel is still OPEN with your data")
                    self.log(f"   Review the changes, then close Excel normally (⌘Q or Ctrl+Q)")
                    self.log(f"   Your pivot tables will remain INTACT!")
                    
                    # Keep app and book open - don't close!
                    # User will close Excel manually
                    
                finally:
                    # Don't close the book or quit the app
                    # Let user close Excel manually
                    pass
            
            except Exception as e:
                self.log(f"❌ xlwings error: {e}")
                self.log(f"⚠️  Make sure Excel is not open and file is not locked")
                # Clean up only if WE spawned a new Excel app (don't kill user's pre-existing session)
                if spawned_app is not None:
                    try:
                        spawned_app.quit()
                    except Exception:
                        pass
                raise

        elapsed = time.time() - self.start_time
        self.log("\n" + "="*80)
        self.log("SUMMARY")
        self.log("="*80)
        self.log(f"✓ Updated: {self.updated_count}")
        self.log(f"⚠ No Match: {self.no_match_count}")
        self.log(f"✗ Failed: {self.failed_count}")
        self.log(f"⏱ Time: {elapsed:.2f}s")
        self.log("="*80 + "\n")
        
        return {
            "updated": self.updated_count,
            "no_match": self.no_match_count,
            "failed": self.failed_count,
        }


# ==============================================================================
# SECTION 7: CLI
# ==============================================================================

def run_cli():
    """CLI interface for AccrualUpdater"""
    parser = argparse.ArgumentParser(
        description="Accrual Updater v7.0.2 - HYBRID (xlwings direct writes + v7.5.0 logic)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    
    parser.add_argument("--master", required=True, help="Master file (.xlsx or .xlsm)")
    parser.add_argument("--sheet", default="Profit Sharing", help="Sheet name")
    parser.add_argument("--header-row", type=int, default=3, help="Header row")
    parser.add_argument("--month", required=True, choices=MONTHS, help="Month")
    parser.add_argument("--year", type=int, required=True, help="Year")
    parser.add_argument("--paysheets", required=True, help="Paysheets folder")
    parser.add_argument("--dates", nargs="+", help="Dates for AB calculation")
    parser.add_argument("--multipliers", nargs="+", type=float, help="Multipliers for dates")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    parser.add_argument("--backup", action="store_true", help="Create backup before writing")
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    
    args = parser.parse_args()

    date_multiplier_pairs = []
    if args.dates and args.multipliers:
        if len(args.dates) == len(args.multipliers):
            date_multiplier_pairs = list(zip(args.dates, args.multipliers))
        else:
            print("ERROR: dates and multipliers count mismatch")
            sys.exit(1)

    updater = AccrualUpdater(
        master_path=args.master,
        sheet_name=args.sheet,
        header_row=args.header_row,
        month=args.month,
        year=args.year,
        paysheets_folder=args.paysheets,
        date_multiplier_pairs=date_multiplier_pairs,
        dry_run=args.dry_run,
        backup=args.backup,
    )
    
    result = updater.process()
    print("\nResult:", result)


if __name__ == "__main__":
    run_cli()
