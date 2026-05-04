"""
iTech Accrual Updater - FastAPI backend
Wraps the existing AccrualUpdater Python logic for web access.
"""
import asyncio
import json
import os
import re
import sys
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# Locate accrual_updater module — works in dev, frozen exe (PyInstaller), and
# any sibling layout. NEVER hardcode user-specific paths.
def _locate_accrual_src() -> Optional[Path]:
    candidates: List[Path] = []
    if getattr(sys, "frozen", False):
        # PyInstaller _MEIPASS bundle
        mei = Path(getattr(sys, "_MEIPASS", ""))
        if mei.exists():
            candidates.append(mei)
            candidates.append(mei / "iTech_Accrual_Updater")
    here = Path(__file__).resolve().parent
    candidates += [
        here,                                  # vendored in backend/
        here.parent / "iTech_Accrual_Updater", # sibling
        here.parent.parent / "iTech_Accrual_Updater",
        here.parent.parent / "ml:ai projects" / "iTech_Accrual_Updater",
    ]
    for c in candidates:
        if (c / "accrual_updater.py").exists():
            return c
    # Last-resort env override
    env = os.environ.get("ACCRUAL_SRC")
    if env and (Path(env) / "accrual_updater.py").exists():
        return Path(env)
    return None


ACCRUAL_SRC = _locate_accrual_src()
if ACCRUAL_SRC and str(ACCRUAL_SRC) not in sys.path:
    sys.path.insert(0, str(ACCRUAL_SRC))
print(f"AccrualUpdater source: {ACCRUAL_SRC or 'NOT FOUND'}")

try:
    from accrual_updater import AccrualUpdater  # type: ignore
    ACCRUAL_AVAILABLE = True
except Exception as e:
    print(f"⚠️  Could not import AccrualUpdater: {e}")
    AccrualUpdater = None  # type: ignore
    ACCRUAL_AVAILABLE = False

try:
    from admin_fee_module_v18b import calculate_admin_fee_for_paysheet  # type: ignore
    ADMIN_FEE_AVAILABLE = True
except Exception:
    ADMIN_FEE_AVAILABLE = False


MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


app = FastAPI(title="iTech Accrual Updater API")

# Allowed roots for filesystem browsing/run. Override with ACCRUAL_ALLOWED_ROOTS
# (colon-sep on Unix, semicolon-sep on Windows — uses os.pathsep).
import platform as _platform
import tempfile as _tempfile
if _platform.system() == "Windows":
    DEFAULT_ROOTS = [
        str(Path.home()),
        "C:\\",  # all of C: (Kristina's drive)
        "D:\\",  # secondary drive if present
        _tempfile.gettempdir(),
    ]
else:
    DEFAULT_ROOTS = [str(Path.home()), _tempfile.gettempdir(), "/Users", "/Volumes"]

_env_roots = os.environ.get("ACCRUAL_ALLOWED_ROOTS", "")
_root_strs = _env_roots.split(os.pathsep) if _env_roots else DEFAULT_ROOTS
ALLOWED_ROOTS: List[Path] = []
for p in _root_strs:
    if not p:
        continue
    try:
        rp = Path(p).resolve()
        # Don't reject roots that don't exist (D:\ may be optional)
        ALLOWED_ROOTS.append(rp)
    except Exception:
        continue


def _safe_path(p: str) -> Path:
    """Resolve user-supplied path. Reject if outside ALLOWED_ROOTS or symlink-escapes.
    Uses Path.relative_to (not str.startswith) to avoid prefix-bypass:
    e.g. '/Users/foo' must NOT match allowed root '/Users/fo'."""
    resolved = Path(p).expanduser().resolve()
    for root in ALLOWED_ROOTS:
        try:
            resolved.relative_to(root)
            return resolved  # path is under or equal to an allowed root
        except ValueError:
            continue
    raise HTTPException(status_code=403, detail=f"Path outside allowed roots: {resolved}")


# DNS-rebinding defense: reject Host headers that aren't loopback
@app.middleware("http")
async def host_header_guard(request, call_next):
    host = (request.headers.get("host") or "").split(":")[0].lower()
    if host not in ("localhost", "127.0.0.1", "::1"):
        from fastapi.responses import JSONResponse
        return JSONResponse({"detail": f"Host {host!r} not allowed"}, status_code=403)
    return await call_next(request)


app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5273", "http://127.0.0.1:5273",
        "http://localhost:5173", "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "Authorization"],
)


# =============================================================================
# Models
# =============================================================================

from pydantic import Field  # noqa: E402


class PayDate(BaseModel):
    date: str = Field(min_length=8, max_length=12)  # MM/DD/YYYY
    multiplier: float = Field(ge=0.0, le=10.0)


class RunRequest(BaseModel):
    master_path: str = Field(min_length=1, max_length=2000)
    paysheets_folder: str = Field(min_length=1, max_length=2000)
    month: str
    year: int = Field(ge=2020, le=2100)
    dry_run: bool = True
    enable_accrual: bool = True
    enable_admin_fee: bool = True
    enable_carryforward: bool = False
    pay_dates: List[PayDate] = Field(default_factory=list, max_length=50)


class BrowseResponse(BaseModel):
    path: str
    parent: Optional[str]
    entries: List[Dict[str, Any]]


class BugReport(BaseModel):
    description: str
    sender_email: Optional[str] = None
    user_agent: Optional[str] = None
    url: Optional[str] = None

    class Config:
        # Size caps to prevent DoS via huge payloads
        str_max_length = 10_000

# Bug-report rate limit (per-process, in-memory)
_BUG_REPORT_TIMES: List[float] = []
_BUG_REPORT_RATE_LIMIT = 10  # max per minute
_BUG_REPORT_WINDOW = 60.0


def _bug_report_rate_check() -> bool:
    import time as _t
    now = _t.time()
    # purge old
    _BUG_REPORT_TIMES[:] = [t for t in _BUG_REPORT_TIMES if now - t < _BUG_REPORT_WINDOW]
    if len(_BUG_REPORT_TIMES) >= _BUG_REPORT_RATE_LIMIT:
        return False
    _BUG_REPORT_TIMES.append(now)
    return True


# =============================================================================
# Job registry - in-memory log queue per job
# =============================================================================

JOBS: Dict[str, Dict[str, Any]] = {}
import threading as _threading
_RUN_LOCK = _threading.Lock()  # Process-wide guard against concurrent runs


def _prune_jobs(max_age_seconds: float = 3600.0) -> None:
    """Drop completed/old jobs to bound memory."""
    now = datetime.utcnow()
    stale = []
    for jid, j in JOBS.items():
        try:
            started = datetime.fromisoformat(j["started_at"])
            if j.get("status") in ("done", "error") and (now - started).total_seconds() > max_age_seconds:
                stale.append(jid)
            elif (now - started).total_seconds() > 6 * 3600:  # hard cap 6h
                stale.append(jid)
        except Exception:
            stale.append(jid)
    for jid in stale:
        JOBS.pop(jid, None)


def _new_job() -> str:
    _prune_jobs()
    job_id = uuid.uuid4().hex
    JOBS[job_id] = {
        "id": job_id,
        "status": "pending",
        "queue": asyncio.Queue(maxsize=10000),  # bound log queue
        "started_at": datetime.utcnow().isoformat(),
    }
    return job_id


async def _push_log(job_id: str, line: str) -> None:
    if job_id in JOBS:
        await JOBS[job_id]["queue"].put(line)


def _push_log_sync(job_id: str, line: str, loop: asyncio.AbstractEventLoop) -> None:
    asyncio.run_coroutine_threadsafe(_push_log(job_id, line), loop)


# =============================================================================
# Endpoints
# =============================================================================

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "accrual_available": ACCRUAL_AVAILABLE,
        "admin_fee_available": ADMIN_FEE_AVAILABLE,
        "months": MONTHS,
    }


@app.get("/api/native-picker")
def native_picker(mode: str = Query(...), title: str = Query(default="Select")):
    """Open OS-native file/folder picker. Returns selected path.
    mode='file' → file picker (Excel only). mode='folder' → directory picker."""
    import platform
    import subprocess
    if mode not in ("file", "folder"):
        raise HTTPException(status_code=400, detail="mode must be 'file' or 'folder'")

    title_clean = re.sub(r'[\x00-\x1f"\\]', '', title)[:120]
    system = platform.system()

    try:
        if system == "Darwin":
            # AppleScript native dialog
            if mode == "file":
                script = (
                    f'tell application "System Events" to activate\n'
                    f'POSIX path of (choose file with prompt "{title_clean}" '
                    f'of type {{"xlsx", "xls", "xlsm"}})'
                )
            else:
                script = (
                    f'tell application "System Events" to activate\n'
                    f'POSIX path of (choose folder with prompt "{title_clean}")'
                )
            result = subprocess.run(
                ["osascript", "-e", script],
                capture_output=True, text=True, timeout=300,
            )
            if result.returncode != 0:
                # User cancelled = exit code 1, stderr empty-ish
                if "User canceled" in result.stderr or result.returncode == 1:
                    return {"path": None, "cancelled": True}
                raise HTTPException(status_code=500, detail=f"Picker error: {result.stderr.strip()}")
            return {"path": result.stdout.strip().rstrip("/"), "cancelled": False}

        if system == "Windows":
            # PowerShell native dialog. Quote title as single-quoted PS literal
            # to avoid expansion of $foo and embedded double-quote injection.
            def _ps_q(s: str) -> str:
                return "'" + s.replace("'", "''") + "'"

            if mode == "file":
                ps = (
                    f'Add-Type -AssemblyName System.Windows.Forms; '
                    f'$d = New-Object System.Windows.Forms.OpenFileDialog; '
                    f'$d.Title = {_ps_q(title_clean)}; '
                    f'$d.Filter = {_ps_q("Excel|*.xlsx;*.xls;*.xlsm")}; '
                    f'if ($d.ShowDialog() -eq "OK") {{ Write-Output $d.FileName }}'
                )
            else:
                ps = (
                    f'Add-Type -AssemblyName System.Windows.Forms; '
                    f'$d = New-Object System.Windows.Forms.FolderBrowserDialog; '
                    f'$d.Description = {_ps_q(title_clean)}; '
                    f'if ($d.ShowDialog() -eq "OK") {{ Write-Output $d.SelectedPath }}'
                )

            # CREATE_NO_WINDOW prevents black console flash; -STA required for OpenFileDialog
            CREATE_NO_WINDOW = 0x08000000
            popen_kwargs: Dict[str, Any] = {
                "capture_output": True,
                "text": True,
                "timeout": 300,
            }
            try:
                popen_kwargs["creationflags"] = CREATE_NO_WINDOW
                si = subprocess.STARTUPINFO()
                si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                popen_kwargs["startupinfo"] = si
            except AttributeError:
                pass

            result = subprocess.run(
                ["powershell", "-NoProfile", "-STA", "-Command", ps],
                **popen_kwargs,
            )
            path = result.stdout.strip()
            if not path:
                return {"path": None, "cancelled": True}
            return {"path": path, "cancelled": False}

        raise HTTPException(status_code=501, detail=f"Native picker not supported on {system}")
    except subprocess.TimeoutExpired:
        return {"path": None, "cancelled": True}


@app.get("/api/browse", response_model=BrowseResponse)
def browse(path: str = Query(default=""), files: bool = Query(default=False)):
    """List directories (and optionally files) at path. If empty, starts at $HOME."""
    if not path:
        path = str(Path.home())
    p = _safe_path(path)
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"Path not found: {p}")
    if not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a directory: {p}")

    entries: List[Dict[str, Any]] = []
    try:
        for child in sorted(p.iterdir(), key=lambda c: (not c.is_dir(), c.name.lower())):
            if child.name.startswith("."):
                continue
            is_dir = child.is_dir()
            if not is_dir and not files:
                continue
            if not is_dir and child.suffix.lower() not in (".xls", ".xlsx", ".xlsm"):
                continue
            entries.append({
                "name": child.name,
                "path": str(child),
                "is_dir": is_dir,
            })
    except PermissionError:
        pass

    parent = str(p.parent) if p.parent != p else None
    return BrowseResponse(path=str(p), parent=parent, entries=entries)


@app.post("/api/run")
async def run(req: RunRequest):
    """Kick off an accrual update job. Returns job_id for log streaming."""
    if not ACCRUAL_AVAILABLE:
        raise HTTPException(status_code=500, detail="AccrualUpdater module not available")

    master_p = _safe_path(req.master_path)
    sheets_p = _safe_path(req.paysheets_folder)
    if not master_p.exists():
        raise HTTPException(status_code=400, detail=f"Master file not found: {master_p}")
    if master_p.suffix.lower() == ".xlsm":
        # .xlsm files can auto-execute macros when opened by xlwings.
        # Require explicit env var to allow.
        if os.environ.get("ACCRUAL_ALLOW_MACROS") != "1":
            raise HTTPException(status_code=403, detail="Macro-enabled .xlsm files blocked. Set ACCRUAL_ALLOW_MACROS=1 to override.")
    if not sheets_p.is_dir():
        raise HTTPException(status_code=400, detail=f"Paysheets folder not found: {sheets_p}")
    if req.month not in MONTHS:
        raise HTTPException(status_code=400, detail=f"Invalid month: {req.month}")
    # Use validated paths
    req.master_path = str(master_p)
    req.paysheets_folder = str(sheets_p)

    # Reject concurrent runs: only one xlwings session at a time
    if not _RUN_LOCK.acquire(blocking=False):
        raise HTTPException(
            status_code=409,
            detail="Another run is already in progress. Wait for it to finish or close the browser tab.",
        )

    job_id = _new_job()
    loop = asyncio.get_running_loop()

    def worker():
        # Initialize COM in this worker thread (Windows xlwings + Outlook)
        co_initialized = False
        if _platform.system() == "Windows":
            try:
                import pythoncom  # type: ignore
                pythoncom.CoInitialize()
                co_initialized = True
            except Exception:
                pass
        try:
            JOBS[job_id]["status"] = "running"
            _push_log_sync(job_id, f"▶ Starting run for {req.month} {req.year}", loop)
            _push_log_sync(job_id, f"  Master: {req.master_path}", loop)
            _push_log_sync(job_id, f"  Paysheets: {req.paysheets_folder}", loop)
            _push_log_sync(job_id, f"  Dry run: {req.dry_run}", loop)
            _push_log_sync(job_id, "", loop)

            if req.enable_accrual:
                _push_log_sync(job_id, "=" * 72, loop)
                _push_log_sync(job_id, "STEP 1: ACCRUAL UPDATE (Hours, Billed, Pay Dates)", loop)
                _push_log_sync(job_id, "=" * 72, loop)

                date_multiplier_pairs = [
                    (pd.date, pd.multiplier) for pd in req.pay_dates
                ]

                updater = AccrualUpdater(
                    master_path=req.master_path,
                    sheet_name="Profit Sharing",
                    header_row=3,
                    month=req.month,
                    year=req.year,
                    paysheets_folder=req.paysheets_folder,
                    date_multiplier_pairs=date_multiplier_pairs,
                    dry_run=req.dry_run,
                    backup=True,
                    enable_carryforward=req.enable_carryforward,
                )
                updater.process()
                for line in updater.log_lines:
                    _push_log_sync(job_id, line, loop)

            if req.enable_admin_fee and ADMIN_FEE_AVAILABLE:
                _push_log_sync(job_id, "", loop)
                _push_log_sync(job_id, "=" * 72, loop)
                _push_log_sync(job_id, "STEP 2: ADMIN FEE CALCULATION", loop)
                _push_log_sync(job_id, "=" * 72, loop)
                _run_admin_fee(req, job_id, loop)

            JOBS[job_id]["status"] = "done"
            _push_log_sync(job_id, "", loop)
            _push_log_sync(job_id, "✓ Run complete.", loop)
            _push_log_sync(job_id, "__DONE__", loop)
        except Exception as e:
            JOBS[job_id]["status"] = "error"
            tb = traceback.format_exc()
            friendly = _humanize_error(e)
            _push_log_sync(job_id, f"✗ ERROR: {friendly}", loop)
            _push_log_sync(job_id, tb, loop)
            _push_log_sync(job_id, "__DONE__", loop)
        finally:
            if co_initialized:
                try:
                    import pythoncom  # type: ignore
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
            _RUN_LOCK.release()

    # Run in a thread to avoid blocking the event loop (xlwings is sync/COM)
    import threading
    threading.Thread(target=worker, daemon=True).start()

    return {"job_id": job_id}


def _humanize_error(e: Exception) -> str:
    """Map technical exceptions to plain English for non-technical users."""
    msg = str(e)
    if isinstance(e, FileNotFoundError) or "No such file" in msg:
        return f"File not found. Check that the master file or paysheets folder still exists. ({msg[:120]})"
    if isinstance(e, KeyError) and "Profit Sharing" in msg:
        return "Sheet 'Profit Sharing' not found in master file. Click Help → Format Rules."
    if "PermissionError" in type(e).__name__ or "Permission denied" in msg:
        return "File is locked or read-only. Close it in Excel and try again."
    if "is not a valid Win32" in msg or "COM" in msg:
        return f"Excel COM error. Restart Excel and try again. ({msg[:120]})"
    if "year" in msg.lower() and "tab" in msg.lower():
        return "No matching year tab found in paysheet. Verify paysheet has a tab named with the year."
    if "Required columns" in msg:
        return "Required columns missing from master. Click Help → Format Rules to see required headers."
    return msg


def _run_admin_fee(req: RunRequest, job_id: str, loop: asyncio.AbstractEventLoop) -> None:
    """STEP 2 admin fee — uses xlwings (preserves queries/external data ranges).
    openpyxl save would strip 'External data range' — known issue.
    """
    import re
    from openpyxl import load_workbook as _opx_load
    from openpyxl.utils import get_column_letter
    import xlwings as xw  # type: ignore
    from accrual_updater import find_headers  # type: ignore

    month_idx = MONTHS.index(req.month) + 1
    HEADER_ROW = 3

    # Read-only pass with openpyxl — just to detect headers + build lookup
    wb_read = _opx_load(req.master_path, read_only=True, data_only=True)
    ws_read = wb_read["Profit Sharing"]
    headers = find_headers(ws_read, HEADER_ROW, req.month)
    file_col = headers.get('file_col') or 1
    admin_col = headers.get('admin_fee_col')

    lookup: Dict[str, int] = {}
    for r in range(HEADER_ROW + 1, ws_read.max_row + 1):
        fn = ws_read.cell(row=r, column=file_col).value
        if fn:
            fn_str = str(fn).strip()
            if 5 <= len(fn_str) <= 6 and fn_str.isdigit():
                lookup[fn_str] = r
    wb_read.close()

    _push_log_sync(job_id, f"Found {len(lookup)} file numbers (col {get_column_letter(file_col)})", loop)

    if not admin_col:
        _push_log_sync(job_id, "❌ Admin Fee column not found", loop)
        return
    _push_log_sync(job_id, f"✓ Admin Fee column: {get_column_letter(admin_col)} ({admin_col})", loop)

    files: List[str] = []
    psheets = Path(req.paysheets_folder)
    for p in psheets.rglob("*.xls"):
        files.append(str(p))
    for p in psheets.rglob("*.xlsx"):
        files.append(str(p))
    files.sort()

    # Compute all admin-fee values first
    to_write: List[Tuple[int, float, str]] = []  # (row, value, fname)
    skipped = 0
    for idx, psheet in enumerate(files, 1):
        fname = os.path.basename(psheet)
        m = re.search(r"(\d{5,6})", fname)
        if not m:
            skipped += 1
            continue
        file_number = m.group(1)
        if file_number not in lookup:
            skipped += 1
            continue
        mrow = lookup[file_number]
        try:
            hours, rate, admin_fee = calculate_admin_fee_for_paysheet(
                paysheet_path=psheet, month=month_idx, year=req.year, debug=False
            )
            if isinstance(admin_fee, (int, float)) and admin_fee > 0:
                to_write.append((mrow, round(float(admin_fee), 2), fname))
                _push_log_sync(job_id, f"[{idx:3d}] {fname}: ${admin_fee:.2f}", loop)
            else:
                skipped += 1
        except Exception as e:
            _push_log_sync(job_id, f"[{idx:3d}] ✗ {fname}: {str(e)[:60]}", loop)

    if req.dry_run or not to_write:
        _push_log_sync(job_id, f"(dry) {len(to_write)} admin fees ready, {skipped} skipped", loop)
        return

    # Write via xlwings — REUSE existing Excel session if STEP 1 left it open.
    # Avoids two-Excel-instance race that strips external data ranges.
    _push_log_sync(job_id, f"Writing {len(to_write)} admin fees via xlwings…", loop)
    target_path = os.path.normcase(os.path.realpath(str(Path(req.master_path).resolve())))
    book = None
    spawned_app = None  # only set if we created a new App
    try:
        # Look for an already-open book matching our target file (any running App)
        for existing_app in list(xw.apps):
            for existing_book in list(existing_app.books):
                try:
                    if os.path.normcase(os.path.realpath(existing_book.fullname)) == os.path.normcase(target_path):
                        book = existing_book
                        _push_log_sync(job_id, f"  ✓ Reusing existing Excel session (PID {existing_app.pid})", loop)
                        break
                except Exception:
                    continue
            if book is not None:
                break

        # If no open match, start fresh
        if book is None:
            spawned_app = xw.App(visible=True)
            book = xw.Book(target_path)
            _push_log_sync(job_id, f"  ✓ Started new Excel session", loop)

        ws_xw = book.sheets["Profit Sharing"]
        written = 0
        for row, value, fname in to_write:
            cur_formula = ws_xw.cells(row, admin_col).formula
            if isinstance(cur_formula, str) and cur_formula.startswith('='):
                _push_log_sync(job_id, f"  ⚠️  {get_column_letter(admin_col)}{row} formula — SKIPPED", loop)
                continue
            ws_xw.cells(row, admin_col).value = value
            written += 1
        book.save()
        _push_log_sync(job_id, f"✅ Saved — {written} admin fees written. Queries/connections preserved!", loop)
        # Leave Excel open for inspection. Do not quit App.
    except Exception as e:
        _push_log_sync(job_id, f"✗ xlwings admin-fee failed: {e}", loop)
        # Cleanup ONLY if we started Excel ourselves and hit error
        if spawned_app is not None:
            try:
                spawned_app.quit()
            except Exception:
                pass


# =============================================================================
# Bug reports
# =============================================================================

BUG_RECIPIENT = os.environ.get("BUG_REPORT_TO", "ravi.vavilala@riseits.com")


def _bug_log_path() -> Path:
    """Pick a writable location for bug-report log.
    Windows: %APPDATA%/AccrualUpdater/. Mac/Linux: ~/Library/Application Support/AccrualUpdater/
    Falls back to next-to-script if env paths fail."""
    try:
        if _platform.system() == "Windows":
            base = os.environ.get("APPDATA")
            if base:
                p = Path(base) / "AccrualUpdater"
                p.mkdir(parents=True, exist_ok=True)
                return p / "bug_reports.jsonl"
        elif _platform.system() == "Darwin":
            p = Path.home() / "Library" / "Application Support" / "AccrualUpdater"
            p.mkdir(parents=True, exist_ok=True)
            return p / "bug_reports.jsonl"
    except Exception:
        pass
    return Path(__file__).parent / "bug_reports.jsonl"


BUG_LOG_FILE = _bug_log_path()


def _send_via_windows_outlook(
    recipient: str, subject: str, body: str, reply_to: Optional[str] = None,
) -> tuple[bool, Optional[str]]:
    """Send via Windows Outlook COM. Sets ReplyRecipients so Ravi can reply
    directly to Kristina's typed email even if Outlook sends as her default."""
    try:
        import win32com.client  # type: ignore
        import pythoncom  # type: ignore
    except ImportError:
        return False, "pywin32 not installed (PyInstaller --hidden-import win32com.client)"
    try:
        # Initialize COM in this thread (uvicorn worker)
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0 = olMailItem
            mail.To = recipient
            mail.Subject = subject
            mail.Body = body
            if reply_to:
                try:
                    mail.ReplyRecipients.Add(reply_to)
                    mail.Save()  # required to persist ReplyRecipients before Send
                except Exception:
                    pass  # if Outlook rejects, still send without Reply-To
            mail.Send()
            return True, "outlook-windows"
        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    except Exception as e:
        return False, f"Windows Outlook COM: {e}"


def _send_via_macos_applescript(
    recipient: str, subject: str, body: str, reply_to: Optional[str] = None,
) -> tuple[bool, Optional[str]]:
    """Send email via Outlook or Mail.app on macOS.
    Body is passed via tempfile to avoid AppleScript string-injection (CVE-class)."""
    import subprocess
    import tempfile

    # Sanitize recipient + subject (used in osascript -e). Reject if suspicious.
    if not re.match(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$', recipient):
        return False, f"Invalid recipient: {recipient!r}"
    # Subject: strip control chars, cap length
    subject = re.sub(r'[\x00-\x1f\x7f]', ' ', subject)[:200]
    # Body via tempfile — never interpolate raw multiline text into AppleScript
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        body_path = f.name
        f.write(body)

    def _q(s: str) -> str:
        # Safe inside AppleScript double-quoted string after stripping controls
        return s.replace("\\", "\\\\").replace('"', '\\"')

    try:
        outlook_script = f'''
        set bodyText to (read POSIX file "{_q(body_path)}" as «class utf8»)
        tell application "Microsoft Outlook"
            set newMessage to make new outgoing message with properties {{subject:"{_q(subject)}", plain text content:bodyText}}
            make new recipient at newMessage with properties {{email address:{{address:"{_q(recipient)}"}}}}
            send newMessage
        end tell
        '''
        result = subprocess.run(
            ["osascript", "-e", outlook_script],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode == 0:
            return True, "outlook-macos"
        outlook_err = result.stderr.strip()

        mail_script = f'''
        set bodyText to (read POSIX file "{_q(body_path)}" as «class utf8»)
        tell application "Mail"
            set newMessage to make new outgoing message with properties {{subject:"{_q(subject)}", content:bodyText, visible:false}}
            tell newMessage
                make new to recipient at end of to recipients with properties {{address:"{_q(recipient)}"}}
                send
            end tell
        end tell
        '''
        result = subprocess.run(
            ["osascript", "-e", mail_script],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode == 0:
            return True, "mail-macos"
        return False, f"Outlook: {outlook_err} | Mail: {result.stderr.strip()}"
    except Exception as e:
        return False, f"Outlook: error | Mail: {e}"
    finally:
        try:
            os.unlink(body_path)
        except Exception:
            pass


def _send_via_native_mail(
    recipient: str, subject: str, body: str, reply_to: Optional[str] = None,
) -> tuple[bool, Optional[str]]:
    """Dispatch to the right native mail sender based on OS."""
    import platform
    system = platform.system()
    if system == "Windows":
        return _send_via_windows_outlook(recipient, subject, body, reply_to=reply_to)
    if system == "Darwin":
        return _send_via_macos_applescript(recipient, subject, body, reply_to=reply_to)
    return False, f"Unsupported OS: {system}"


@app.post("/api/bug-report")
def bug_report(report: BugReport):
    """Log bug report locally and send email via SMTP or macOS AppleScript."""
    if not _bug_report_rate_check():
        raise HTTPException(status_code=429, detail="Too many bug reports. Try again in a minute.")

    # Cap field sizes (defensive — pydantic str_max_length is also set)
    desc = (report.description or "")[:10_000].strip()
    sender = (report.sender_email or "")[:200] or None
    url = (report.url or "")[:500] or None
    ua = (report.user_agent or "")[:500] or None
    if not desc:
        raise HTTPException(status_code=400, detail="Description required")

    record = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "recipient": BUG_RECIPIENT,
        "description": desc,
        "sender_email": sender,
        "user_agent": ua,
        "url": url,
    }

    # Always persist locally — rotate if file exceeds 10 MB
    try:
        if BUG_LOG_FILE.exists() and BUG_LOG_FILE.stat().st_size > 10 * 1024 * 1024:
            BUG_LOG_FILE.rename(str(BUG_LOG_FILE) + ".old")
        with open(BUG_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
    except Exception as e:
        print(f"⚠️  Could not write bug log: {e}")

    # Override with sanitized values for downstream email
    report.description = desc
    report.sender_email = sender
    report.url = url
    report.user_agent = ua

    subject = "7t.ai Accrual Updater — Bug Report"
    body = (
        f"Time (UTC): {record['timestamp']}\n"
        f"From:       {report.sender_email or 'anonymous'}\n"
        f"URL:        {report.url or '-'}\n"
        f"User-Agent: {report.user_agent or '-'}\n"
        f"\n"
        f"Description:\n"
        f"{report.description.strip()}\n"
    )

    # Try SMTP first if configured
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    sent = False
    via: Optional[str] = None
    error: Optional[str] = None

    if smtp_host and smtp_user and smtp_pass:
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            smtp_port = int(os.environ.get("SMTP_PORT", "587"))
            msg = MIMEMultipart()
            msg["From"] = smtp_user
            msg["To"] = BUG_RECIPIENT
            if report.sender_email:
                msg["Reply-To"] = report.sender_email
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, [BUG_RECIPIENT], msg.as_string())
            sent = True
            via = "smtp"
        except Exception as e:
            error = f"SMTP: {e}"
            print(f"⚠️  SMTP send failed: {e}")

    # Fall back to native mail (Windows Outlook COM / macOS AppleScript)
    if not sent:
        ok, result = _send_via_native_mail(BUG_RECIPIENT, subject, body, reply_to=report.sender_email)
        if ok:
            sent = True
            via = result
        else:
            error = f"{error + ' | ' if error else ''}Native: {result}"

    return {
        "ok": True,
        "logged_to": str(BUG_LOG_FILE),
        "emailed_to": BUG_RECIPIENT if sent else None,
        "via": via,
        "error": error,
    }


@app.get("/api/logs/{job_id}")
async def stream_logs(job_id: str):
    """SSE stream of log lines for a job."""
    if job_id not in JOBS:
        raise HTTPException(status_code=404, detail="Job not found")

    queue: asyncio.Queue = JOBS[job_id]["queue"]

    async def event_gen():
        while True:
            line = await queue.get()
            if line == "__DONE__":
                yield f"event: done\ndata: {json.dumps({'status': JOBS[job_id]['status']})}\n\n"
                break
            yield f"data: {json.dumps({'line': line})}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream")


# =============================================================================
# Serve built React app at /
# =============================================================================

DIST_DIR = Path(__file__).parent.parent / "frontend" / "dist"
if DIST_DIR.exists():
    app.mount("/assets", StaticFiles(directory=DIST_DIR / "assets"), name="assets")

    @app.get("/{full_path:path}")
    def spa(full_path: str):
        # Never hijack API routes — let FastAPI return proper 404/405
        if full_path.startswith("api/") or full_path.startswith("api"):
            raise HTTPException(status_code=404, detail="Not Found")
        index = DIST_DIR / "index.html"
        if index.exists():
            return FileResponse(index)
        return {"error": "frontend not built"}
